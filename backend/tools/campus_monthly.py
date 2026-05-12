"""校内月度分析 — 基于签到累积数据 + 财务/课程类型/退款上传文件"""
import os
import re
import io
import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import json

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")
UPLOAD_DIR = os.path.join(DATA_DIR, "monthly_output")

def _read_excel(data: bytes, sheet_name: str) -> pd.DataFrame:
    """自动选择 Excel 引擎：先尝试 openpyxl（.xlsx），失败后退回 xlrd（.xls）"""
    try:
        return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine='openpyxl')
    except Exception:
        return pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, engine='xlrd')


def _get_dept_file() -> str:
    config_path = os.path.join(DATA_DIR, "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    rel = config["tools"]["department_mapping"]
    return os.path.join(DATA_DIR, "..", rel)


# ──────────────────── 工具函数 ────────────────────

def extract_school_main(school_name):
    """提取学校主名称（归总校区）"""
    if pd.isna(school_name):
        return '未知学校'
    match = re.match(r'^(.+?)(?:（[^）]+）|$)', str(school_name))
    return match.group(1).strip() if match else str(school_name).strip()


def auto_column_width(ws):
    """中文自适应列宽"""
    def char_width(text):
        if pd.isna(text) or text == '':
            return 2
        return sum(2 if '一' <= c <= '鿿' else 1 for c in str(text))
    for col in ws.columns:
        max_w = max(char_width(cell.value) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 2, 40)


# ──────────────────── 分析函数 ────────────────────

def coach_checkin_analysis(att_df: pd.DataFrame) -> pd.DataFrame:
    """教练考勤分析：签到状态统计 + 签到率"""
    att_df = att_df.copy()
    att_df['签到状态'] = att_df['签到状态'].astype(str).str.strip().fillna('未签到')

    # 签到状态透视
    checkin_stats = (
        att_df.groupby(['部门', '学校名称', '教练姓名', '签到状态'])
        .size()
        .unstack(fill_value=0)
        .add_suffix('次数')
    )
    checkin_stats['总计次数'] = checkin_stats.sum(axis=1)
    checkin_stats['在岗次数'] = checkin_stats.get('在岗次数', 0)
    checkin_stats['教练个人签到率(%)'] = (
        (checkin_stats['在岗次数'] / checkin_stats['总计次数'] * 100)
        .round(2).fillna(0)
    )

    # 学校签到率
    att_df['学校主名称'] = att_df['学校名称'].apply(extract_school_main)
    school_rate = att_df.groupby(['部门', '学校主名称']).apply(
        lambda x: (x['签到状态'] == '在岗').sum() / len(x) * 100 if len(x) > 0 else 0
    ).round(2).rename('学校签到率(%)')

    # 部门签到率
    dept_rate = att_df.groupby('部门').apply(
        lambda x: (x['签到状态'] == '在岗').sum() / len(x) * 100 if len(x) > 0 else 0
    ).round(2).rename('部门签到率(%)')

    result = checkin_stats.reset_index()
    result['学校主名称'] = result['学校名称'].apply(extract_school_main)
    result = result.merge(school_rate.reset_index(), on=['部门', '学校主名称'], how='left')
    result = result.merge(dept_rate.reset_index(), on='部门', how='left')

    # 签到率仅在每组首行显示
    result['学校签到率(%)'] = result.groupby(['部门', '学校主名称'])['学校签到率(%)'].transform(
        lambda x: x.iloc[0] if len(x) > 0 else ''
    ).where(result.groupby(['部门', '学校主名称']).cumcount() == 0, '')

    result['部门签到率(%)'] = result.groupby('部门')['部门签到率(%)'].transform(
        lambda x: x.iloc[0] if len(x) > 0 else ''
    ).where(result.groupby('部门').cumcount() == 0, '')

    result = result.set_index(['部门', '学校名称', '教练姓名']).drop('学校主名称', axis=1)

    # 添加总计行
    count_cols = [col for col in result.columns if '次数' in col]
    total_row_data = {col: result[col].sum() if col in count_cols else '-' for col in result.columns}
    total_row = pd.Series(total_row_data, name=('总计', '总计', '总计'))
    final_result = pd.concat([result, total_row.to_frame().T])

    # 格式化签到率
    mask = final_result.index != ('总计', '总计', '总计')
    for col in ['教练个人签到率(%)', '学校签到率(%)', '部门签到率(%)']:
        final_result.loc[mask, col] = final_result.loc[mask, col].apply(
            lambda x: f"{x:.2f}%" if x != '' and pd.notna(x) and isinstance(x, (int, float)) else ''
        )

    return final_result


def campus_finance_analysis(coach_att: pd.DataFrame, finance_df: pd.DataFrame) -> pd.DataFrame:
    """校内财务分析：上课人次 + 收入统计"""
    finance_df = finance_df[finance_df['学员类型'] == '学校学员'].copy()

    for col in ['教练', '课程名称']:
        finance_df[col] = finance_df[col].astype(str).str.strip().fillna('未知')

    str_cols = ['教练姓名', '课程名称', '部门', '学校名称']
    for col in str_cols:
        coach_att[col] = coach_att[col].astype(str).str.strip().fillna('未知')

    finance_df['名加课'] = finance_df['教练'] + '_' + finance_df['课程名称']
    coach_att['名加课'] = coach_att['教练姓名'] + '_' + coach_att['课程名称']

    finance_agg = finance_df.groupby('名加课').agg(
        上课人次=('上课日期', 'count'),
        已确认收入=('课程单价', lambda x: pd.to_numeric(x, errors='coerce').fillna(0).sum())
    ).round(2)

    coach_agg = coach_att.groupby(['部门', '学校名称', '教练姓名', '名加课']).size().rename('上课次数').reset_index()

    result = pd.merge(coach_agg, finance_agg.reset_index(), on='名加课', how='left').fillna(0)

    pivot = result.pivot_table(
        index=['部门', '学校名称', '名加课', '教练姓名'],
        values=['上课次数', '上课人次', '已确认收入'],
        aggfunc='sum', fill_value=0, margins=True, margins_name='总计'
    )
    pivot[['上课次数', '上课人次']] = pivot[['上课次数', '上课人次']].astype(int)
    return pivot


def _get_target_types() -> list:
    config_path = os.path.join(DATA_DIR, "config.json")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
        return config.get("tools", {}).get("target_course_types", ['兴趣班', '校队', '梯队'])
    except Exception:
        return ['兴趣班', '校队', '梯队']


def course_type_analysis(att_df: pd.DataFrame, type_file: pd.DataFrame, year: int = 0, month: int = 0) -> Optional[pd.DataFrame]:
    """课程类型分析：课次统计 + 收费类占比"""
    att_df = att_df.copy()
    type_file = type_file.copy()

    att_df.columns = att_df.columns.str.strip()
    type_file.columns = type_file.columns.str.strip()

    if '课程名称' not in att_df.columns or '课程名称' not in type_file.columns:
        return None

    att_df['课程名称'] = att_df['课程名称'].astype(str).str.strip()
    type_file['课程名称'] = type_file['课程名称'].astype(str).str.strip()

    merged_df = pd.merge(att_df, type_file, on='课程名称', how='left')

    missing_type_mask = pd.isna(merged_df['类型'])
    if missing_type_mask.any():
        missing_courses = merged_df.loc[missing_type_mask, '课程名称'].unique().tolist()
        raise ValueError(f"以下课程未配置类型: {', '.join(missing_courses[:10])}")

    if '部门' not in merged_df.columns:
        return None

    att_df_type_ana = merged_df.groupby(['部门', '类型']).size().rename('课程类型次数').reset_index()
    att_df_type_ana_total = att_df_type_ana.groupby('部门')['课程类型次数'].sum().rename('总课次').reset_index()

    target_types = _get_target_types()
    month_str = f"{month}月" if month else ''
    output_list = []

    if not att_df_type_ana_total.empty:
        for _, dept_row in att_df_type_ana_total.iterrows():
            dept_name = dept_row['部门']
            total_count = dept_row['总课次']
            dept_detail = att_df_type_ana[att_df_type_ana['部门'] == dept_name]

            type_str_list = []
            target_total = 0
            for _, type_row in dept_detail.iterrows():
                type_name = str(type_row['类型']) if pd.notna(type_row['类型']) else '未知类型'
                type_count = type_row['课程类型次数']
                type_str_list.append(f'{type_name}{type_count}次')
                if type_name in target_types:
                    target_total += type_count

            target_ratio = (target_total / total_count) * 100 if total_count != 0 else 0
            type_str = '、'.join(type_str_list)
            target_names = '、'.join(target_types)
            output = f"{dept_name}：{month_str}总课次{total_count}次，其中{type_str}，（收费类{target_names}）课次占总课次的{target_ratio:.2f}%"

            output_list.append({
                '部门': dept_name, '总课次': total_count,
                '各类型课次': type_str, '收费类课次': target_total,
                '收费类占比(%)': round(target_ratio, 2), '输出文本': output,
            })

    output_df = pd.DataFrame(output_list)
    final_df = pd.merge(
        att_df_type_ana,
        output_df[['部门', '总课次', '收费类课次', '收费类占比(%)', '输出文本']],
        on='部门', how='left'
    )
    return final_df


def refund_analysis(refund_df: pd.DataFrame, department: pd.DataFrame) -> pd.DataFrame:
    """退款分析"""
    refund_df = pd.merge(refund_df, department, on='学校', how='left')
    pivot = pd.pivot_table(
        refund_df,
        index=['部门', '学校'],
        values=['学员姓名', '退款金额'],
        aggfunc={'学员姓名': 'count', '退款金额': 'sum'},
        margins=True, margins_name='总计'
    )
    return pivot


# ──────────────────── Excel 导出 ────────────────────

def _generate_excel(data_dict: Dict[str, pd.DataFrame]) -> bytes:
    """生成多 sheet Excel 文件，返回字节流"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for sheet_name, df in data_dict.items():
            if df is None:
                continue
            index_flag = isinstance(df.index, pd.MultiIndex) or (
                hasattr(df.index, 'nlevels') and df.index.nlevels > 1
            )
            df.to_excel(writer, sheet_name=sheet_name, index=index_flag)

            ws = writer.sheets[sheet_name]
            center = Alignment(horizontal='center', vertical='center')
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center
            auto_column_width(ws)

    buf.seek(0)
    return buf.read()


# ──────────────────── 核心入口 ────────────────────

def preview_monthly(year: int, month: int, checkin_records: List[Dict]) -> Dict[str, Any]:
    """预览：返回该月签到数据概览"""
    if not checkin_records:
        return {"checkin_count": 0, "departments": [], "coaches": [], "schools": []}

    df = pd.DataFrame(checkin_records)
    return {
        "checkin_count": len(df),
        "departments": sorted(df['department'].dropna().unique().tolist()),
        "coaches": sorted(df['coach_name'].dropna().unique().tolist()),
        "schools": sorted(df['school_name'].dropna().unique().tolist()),
        "date_range": [
            str(df['check_date'].min()) if 'check_date' in df.columns else '',
            str(df['check_date'].max()) if 'check_date' in df.columns else '',
        ],
    }


def process_monthly(
    year: int,
    month: int,
    checkin_records: List[Dict],
    finance_bytes: Optional[bytes] = None,
    course_type_bytes: Optional[bytes] = None,
    refund_bytes: Optional[bytes] = None,
) -> Dict[str, Any]:
    """处理月度分析，返回结果 + Excel 字节"""
    if not checkin_records:
        raise ValueError(f"{year}年{month}月无签到数据，请先使用每日签到工具积累数据")

    # 将签到记录转为 DataFrame，统一列名给分析函数用
    att_df = pd.DataFrame(checkin_records)
    att_df = att_df.rename(columns={
        'department': '部门',
        'school_name': '学校名称',
        'course_type': '课程类型',
        'course_name': '课程名称',
        'coach_name': '教练姓名',
        'course_date': '课程日期',
        'sign_in_time': '签到时间',
        'sign_out_time': '签退时间',
        'sign_status': '签到状态',
        'actual_count': '实际上课人次',
        'expected_count': '课程应到人次',
        'confirmed_revenue': '确认收入',
    })

    data_to_export = {}

    # 1. 教练考勤（必做，纯签到数据）
    data_to_export['教练考勤'] = coach_checkin_analysis(att_df)

    # 2. 教练签到原始数据
    data_to_export['教练签到'] = att_df

    # 3. 校内财务分析（需上传财务文件）
    if finance_bytes:
        finance_df = _read_excel(finance_bytes, sheet_name='财务')
        data_to_export['校内分析'] = campus_finance_analysis(att_df.copy(), finance_df)

    # 4. 课程类型分析（需上传课程类型文件）
    if course_type_bytes:
        type_file = _read_excel(course_type_bytes, sheet_name='Sheet1')
        data_to_export['类型分析'] = course_type_analysis(att_df.copy(), type_file, year, month)

    # 5. 退款分析（需上传退款文件 + 部门划分）
    if refund_bytes:
        refund_df = _read_excel(refund_bytes, sheet_name='导出')
        dept_file = _get_dept_file()
        if os.path.exists(dept_file):
            department = pd.read_excel(dept_file, sheet_name='Sheet1')
            data_to_export['退款分析'] = refund_analysis(refund_df, department)

    # 生成 Excel
    excel_bytes = _generate_excel(data_to_export)

    # 保存到磁盘
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filename = f"{year}年{month}月校内分析.xlsx"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, 'wb') as f:
        f.write(excel_bytes)

    return {
        "filename": filename,
        "filepath": filepath,
        "sheets": list(data_to_export.keys()),
        "record_count": len(checkin_records),
        "department_count": len(att_df['部门'].unique()),
        "coach_count": len(att_df['教练姓名'].unique()),
    }
