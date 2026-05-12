"""签到数据整合 — 月度签到数据校对、补签、确认"""
import os
import io
import json
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")
CONSOLIDATION_DIR = os.path.join(DATA_DIR, "consolidation_output")


def load_month_records(records: List[Dict]) -> List[Dict[str, Any]]:
    """Add _row_id to each record for frontend tracking."""
    result = []
    for i, r in enumerate(records):
        row = dict(r)
        row["_row_id"] = i + 1
        result.append(row)
    return result


def save_consolidation_data(year: int, month: int, records: List[Dict]) -> Dict[str, Any]:
    """Save consolidated records as JSON to disk."""
    os.makedirs(CONSOLIDATION_DIR, exist_ok=True)
    filename = f"{year}年{month}月签到整合.json"
    filepath = os.path.join(CONSOLIDATION_DIR, filename)

    # Strip internal _row_id before saving
    clean_records = []
    for r in records:
        rec = {k: v for k, v in r.items() if k != "_row_id"}
        clean_records.append(rec)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(clean_records, f, ensure_ascii=False, indent=2)

    return {"filename": filename, "record_count": len(clean_records)}


def load_consolidation_data(filename: str) -> List[Dict[str, Any]]:
    """Load consolidated records from JSON file."""
    filepath = os.path.join(CONSOLIDATION_DIR, filename)
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def export_excel(records: List[Dict]) -> bytes:
    """Generate Excel from consolidated records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "签到整合"

    col_map = [
        ("check_date", "日期"),
        ("department", "部门"),
        ("school_name", "学校"),
        ("course_type", "课程类型"),
        ("course_name", "课程名称"),
        ("coach_name", "教练"),
        ("course_date", "课程日期"),
        ("start_time", "开始时间"),
        ("end_time", "结束时间"),
        ("sign_in_time", "签到时间"),
        ("sign_out_time", "签退时间"),
        ("sign_status", "签到状态"),
        ("actual_count", "实际上课人次"),
        ("expected_count", "课程应到人次"),
        ("confirmed_revenue", "确认收入"),
        ("remark", "备注"),
    ]

    headers = [h for _, h in col_map]
    ws.append(headers)

    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.alignment = center

    for r in records:
        row = [r.get(key, "") for key, _ in col_map]
        ws.append(row)

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center

    # Auto column width
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell_value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
