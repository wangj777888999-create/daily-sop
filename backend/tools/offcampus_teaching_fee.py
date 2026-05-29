"""校外课时费计算

数据来源：外部数据库（通过 database_manager 执行 SQL）
- 上课记录：字段包含 课程名称、课程信息（含上课日期）、状态（已到/请假/旷课…）、创建时间
- 财务明细：字段包含 课程名称、上课日期、教练、订单类型（校区）、course_pack_name、service_amount

计算逻辑：
1. 新开班配置（独立 JSON）：指定课程名称 + 开课日期，开课日期后 57 天自动结束
2. 将上课记录与财务明细按 (课程名称 + 上课日期) JOIN，获取每条上课记录的归属教练
3. 排除新开班记录后，按教练统计到课率 (已到 / 应到)
4. 根据到课率区间确定「正常课包」阶梯单价
5. 课时费规则：
   - 超时（任何课包）→ 0 元，保留明细
   - 正常课包 / 新开班  → 阶梯单价（按教练当月到课率区间）
   - 跃动课包           → 固定单价（默认 10 元/人次）
   - 168 拼团 / 168 课包 → 0 元，保留明细
   - 一对一私教         → service_amount × 比率（默认 12%）
6. 按教练聚合 → coach_results 结构（含到课率 + 分类明细 + 合计）
"""
import io
import json
import logging
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")
RULES_PATH = os.path.join(DATA_DIR, "offcampus_fee_rules.json")
NEW_CLASS_RULES_PATH = os.path.join(DATA_DIR, "offcampus_new_class_rules.json")
OUTPUT_DIR = os.path.join(DATA_DIR, "offcampus_fee_output")

# ──────────────────── 默认规则 ────────────────────

DEFAULT_RULES: Dict[str, Any] = {
    "sport_keywords": {
        "football": ["足球"],
        "basketball": ["篮球", "农都城"],
    },
    "package_keywords": {
        "football": {
            "跃动课包":    ["跃动"],
            "168拼团课包": ["168"],
            "正常课包":    [],
        },
        "basketball": {
            "168课包":    ["168"],
            "一对一私教": ["私教", "一对一"],
            "正常课包":   [],
        },
    },
    # 到课率阶梯（降序），每档 football/basketball 为正常课包单价（元/人次）
    "attendance_tiers": [
        {"min": 0.9, "football": 18, "basketball": 17},
        {"min": 0.8, "football": 17, "basketball": 16},
        {"min": 0.7, "football": 16, "basketball": 15},
        {"min": 0.5, "football": 10, "basketball": 10},
        {"min": 0.0, "football":  5, "basketball":  5},
    ],
    # 特殊课包固定参数
    "special_fees": {
        "yuedong_per_session": 10,
        "one_on_one_rate":     0.12,
    },
    "course_pack_mapping": {},
    "overtime_threshold_days": 3,
    "coach_fees": {},
    # 保留 fees 结构兼容旧数据（当前计算不再使用）
    "fees": {
        "football": {
            "正常课包":    {"准时": 0, "超时": 0},
            "跃动课包":    {"准时": 10, "超时": 0},
            "168拼团课包": {"准时": 0,  "超时": 0},
        },
        "basketball": {
            "正常课包":   {"准时": 0, "超时": 0},
            "168课包":    {"准时": 0, "超时": 0},
            "一对一私教": {"准时": 0, "超时": 0},
        },
    },
}

SPORT_LABELS = {"football": "足球", "basketball": "篮球"}


# ──────────────────── 规则存取 ────────────────────

def load_fee_rules() -> Dict[str, Any]:
    try:
        with open(RULES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return DEFAULT_RULES
    except Exception as e:
        logger.warning(f"读取费率规则失败，使用默认值: {e}")
        return DEFAULT_RULES


def save_fee_rules(rules: Dict[str, Any]) -> None:
    with open(RULES_PATH, "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)


def load_new_class_rules() -> Dict[str, Any]:
    try:
        with open(NEW_CLASS_RULES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {"new_classes": []}
    except Exception as e:
        logger.warning(f"读取新开班规则失败: {e}")
        return {"new_classes": []}


def save_new_class_rules(rules: Dict[str, Any]) -> None:
    with open(NEW_CLASS_RULES_PATH, "w", encoding="utf-8") as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)


# ──────────────────── 数据获取 ────────────────────

def fetch_data(class_sql: str, finance_sql: str) -> tuple:
    from database_manager import execute_query

    class_result = execute_query(class_sql, limit=50000)
    if not class_result["columns"]:
        raise ValueError("上课记录查询返回空结果，请检查 SQL")
    class_df = pd.DataFrame(class_result["rows"], columns=class_result["columns"])

    finance_result = execute_query(finance_sql, limit=50000)
    if not finance_result["columns"]:
        raise ValueError("财务明细查询返回空结果，请检查 SQL")
    finance_df = pd.DataFrame(finance_result["rows"], columns=finance_result["columns"])

    preview = {
        "class_records": {
            "columns": class_result["columns"],
            "rows": class_result["rows"][:10],
            "total_rows": len(class_df),
        },
        "finance_records": {
            "columns": finance_result["columns"],
            "rows": finance_result["rows"][:10],
            "total_rows": len(finance_df),
        },
    }
    return class_df, finance_df, preview


# ──────────────────── 识别工具 ────────────────────

def detect_sport(course_name: str, sport_keywords: Dict[str, List[str]]) -> str:
    for sport, kws in sport_keywords.items():
        if any(kw in course_name for kw in kws):
            return sport
    return "unknown"


def detect_package_type(course_name: str, sport: str, package_keywords: Dict) -> str:
    sport_pkgs = package_keywords.get(sport, {})
    for pkg_type, kws in sport_pkgs.items():
        if not kws:
            continue
        if any(kw in course_name for kw in kws):
            return pkg_type
    for pkg_type, kws in sport_pkgs.items():
        if not kws:
            return pkg_type
    return "正常课包"


def _extract_class_date(上课信息: Any) -> Optional[str]:
    if not 上课信息:
        return None
    m = re.match(r"(\d{4}-\d{2}-\d{2})", str(上课信息).strip())
    return m.group(1) if m else None


def _parse_date(s: Any) -> Optional[datetime]:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _find_col(col_map: Dict[str, str], candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in col_map:
            return col_map[c]
    return None


# ──────────────────── 核心计算 ────────────────────

def calculate_teaching_fees(
    class_df: pd.DataFrame,
    finance_df: pd.DataFrame,
    rules: Dict[str, Any],
    new_class_config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    到课率阶梯式课时费计算。

    到课率计算范围：排除新开班期间课次后，按教练统计 已到 / 应到。
    新开班期间课次：计入课时费（使用同教练相同阶梯），不计入到课率。
    """
    sport_keywords   = rules.get("sport_keywords",   DEFAULT_RULES["sport_keywords"])
    package_keywords = rules.get("package_keywords", DEFAULT_RULES["package_keywords"])
    course_pack_mapping: Dict[str, Any] = rules.get("course_pack_mapping", {})
    attendance_tiers: List[Dict] = rules.get("attendance_tiers", DEFAULT_RULES["attendance_tiers"])
    special_fees: Dict = rules.get("special_fees", DEFAULT_RULES.get("special_fees", {}))
    threshold = int(rules.get("overtime_threshold_days", 2))

    yuedong_fee     = float(special_fees.get("yuedong_per_session", 10))
    one_on_one_rate = float(special_fees.get("one_on_one_rate", 0.12))

    # 排除配置：新开班（按课程+日期）+ 关键词（如"假期"）
    cfg = new_class_config or {}
    new_classes = cfg.get("new_classes", [])
    exclusion_keywords: List[str] = [
        str(kw).strip() for kw in cfg.get("exclusion_keywords", ["假期"]) if str(kw).strip()
    ]
    exclusion_dates: set = {
        str(d).strip() for d in cfg.get("exclusion_dates", []) if str(d).strip()
    }

    new_class_map: Dict[str, tuple] = {}
    for nc in new_classes:
        cn = str(nc.get("course_name", "")).strip()
        sd = str(nc.get("start_date", "")).strip()
        ed = str(nc.get("end_date", "")).strip()
        if cn:
            new_class_map[cn] = (sd, ed)

    def _is_excluded_from_attendance(course_name: str, class_date: str) -> bool:
        """以下情况不计入到课率：
        1. 课程名含排除关键词（如假期）
        2. 上课日期在指定放假日期列表中
        3. 新开班日期范围内
        """
        # 1. 关键词排除
        for kw in exclusion_keywords:
            if kw in course_name:
                return True
        # 2. 指定放假日期排除
        if str(class_date) in exclusion_dates:
            return True
        # 3. 新开班日期范围排除
        if course_name not in new_class_map:
            return False
        sd, ed = new_class_map[course_name]
        if not sd:
            return True
        return sd <= str(class_date) <= ed

    # ── 1. 预处理上课记录 ──
    class_df = class_df.copy()
    class_cols = {c.strip(): c for c in class_df.columns}

    col_cls_course  = _find_col(class_cols, ["课程名称"])
    col_cls_info    = _find_col(class_cols, ["课程信息", "上课信息", "info"])
    col_cls_status  = _find_col(class_cols, ["状态", "status_type"])
    col_cls_created = _find_col(class_cols, ["创建时间", "create_time"])
    col_cls_member  = _find_col(class_cols, ["会员姓名", "member_name", "学员姓名", "student_name", "姓名"])

    if not col_cls_course:
        raise ValueError(f"上课记录缺少「课程名称」字段，实际列: {list(class_df.columns)}")

    class_df["_course"] = class_df[col_cls_course].astype(str).str.strip()
    class_df["_class_date"] = (
        class_df[col_cls_info].apply(_extract_class_date) if col_cls_info else pd.Series([None] * len(class_df))
    )
    class_df["_status"]   = class_df[col_cls_status].astype(str).str.strip() if col_cls_status else "未知"
    class_df["_attended"] = class_df["_status"] == "已到"
    if col_cls_member:
        class_df["_member"] = class_df[col_cls_member].astype(str).str.strip()

    if col_cls_created:
        class_df["_created_dt"] = class_df[col_cls_created].apply(_parse_date)

    class_df = class_df.dropna(subset=["_class_date"])

    # ── 2. 预处理财务明细 ──
    finance_df = finance_df.copy()
    fin_cols = {c.strip(): c for c in finance_df.columns}

    col_f_course  = _find_col(fin_cols, ["课程名称", "course_name"])
    col_f_date    = _find_col(fin_cols, ["上课日期", "class_date"])
    col_coach     = _find_col(fin_cols, ["教练", "教练姓名", "coach_name", "coach"])
    col_pack      = _find_col(fin_cols, ["course_pack_name", "课包名称"])
    col_service   = _find_col(fin_cols, ["unit_price"])
    col_campus    = _find_col(fin_cols, ["订单类型", "order_type", "校区", "campus"])
    col_quality   = _find_col(fin_cols, ["师资", "教练级别", "quality_type", "coach_level"])
    col_f_member  = _find_col(fin_cols, ["会员姓名", "member_name", "学员姓名", "student_name", "姓名"])

    if not all([col_f_course, col_f_date, col_coach]):
        missing = [n for n, c in [("课程名称", col_f_course), ("上课日期", col_f_date), ("教练", col_coach)] if not c]
        raise ValueError(f"财务明细缺少必要字段: {missing}，实际列: {list(finance_df.columns)}")

    finance_df["_course"]  = finance_df[col_f_course].astype(str).str.strip()
    finance_df["_class_date"] = (
        pd.to_datetime(finance_df[col_f_date], errors="coerce").dt.strftime("%Y-%m-%d")
    )
    finance_df["_coach"] = finance_df[col_coach].astype(str).str.strip()

    if col_pack:
        finance_df["_pack_name"] = (
            finance_df[col_pack].fillna("").astype(str).str.strip()
            .replace("None", "").replace("nan", "")
        )
    else:
        finance_df["_pack_name"] = ""

    finance_df["_unit_price"] = (
        pd.to_numeric(finance_df[col_service], errors="coerce").fillna(0.0) if col_service else 0.0
    )
    finance_df["_campus"]  = finance_df[col_campus].astype(str).str.strip() if col_campus else ""
    finance_df["_quality"] = finance_df[col_quality].astype(str).str.strip() if col_quality else ""
    if col_f_member:
        finance_df["_member"] = finance_df[col_f_member].astype(str).str.strip()
    finance_df = finance_df.dropna(subset=["_class_date"])

    # ── 3. 超时判定 ──
    # 规则：状态为「已到」的记录，创建日期 − 上课日期 ≥ threshold 天 → 超时
    # 超时粒度：
    #   - 两张表均含会员姓名 → 按学员精确匹配（只有该学员超时，同节课其他学员不受影响）
    #   - 缺少会员姓名字段  → 退回整课超时（兼容旧逻辑）
    overtime_set: set = set()               # (course, date) 级别，用于明细展示 & 教练关联
    overtime_student_set: Optional[set] = None  # (course, date, member) 级别，用于费用计算
    overtime_details: List[Dict] = []
    if col_cls_created and "_created_dt" in class_df.columns:
        attended_ot = class_df[
            class_df["_attended"] & class_df["_created_dt"].notna()
        ].copy()

        if len(attended_ot) > 0:
            attended_ot["_class_date_dt"] = pd.to_datetime(attended_ot["_class_date"])
            attended_ot["_days_late"] = attended_ot.apply(
                lambda r: (r["_created_dt"].date() - r["_class_date_dt"].date()).days,
                axis=1,
            )
            ot_records = attended_ot[attended_ot["_days_late"] >= threshold]

            # 课次级别集合（用于展示明细 & 补充教练信息）
            overtime_set = set(zip(ot_records["_course"], ot_records["_class_date"]))

            # 学员级别集合（两张表都有会员列时才启用精确匹配）
            can_match_student = (
                col_cls_member is not None
                and col_f_member is not None
                and "_member" in ot_records.columns
            )
            if can_match_student:
                overtime_student_set = set(
                    zip(ot_records["_course"], ot_records["_class_date"], ot_records["_member"])
                )

            # 明细：按 (课程, 日期) 聚合，late_count = 超时学员去重数
            if len(ot_records) > 0:
                if can_match_student:
                    ot_agg = (
                        ot_records.groupby(["_course", "_class_date"])
                        .agg(
                            _max_created=("_created_dt", "max"),
                            _late_count=("_member",     "nunique"),  # 去重学员数
                            _max_days=("_days_late",    "max"),
                        )
                        .reset_index()
                    )
                else:
                    ot_agg = (
                        ot_records.groupby(["_course", "_class_date"])
                        .agg(
                            _max_created=("_created_dt", "max"),
                            _late_count=("_days_late",   "count"),
                            _max_days=("_days_late",     "max"),
                        )
                        .reset_index()
                    )
                overtime_details = [
                    {
                        "course":         str(row["_course"]),
                        "date":           str(row["_class_date"]),
                        "latest_created": str(row["_max_created"].date()),
                        "max_days_late":  int(row["_max_days"]),
                        "late_count":     int(row["_late_count"]),
                    }
                    for _, row in ot_agg.sort_values("_class_date").iterrows()
                ]

    # ── 4. 将上课记录与教练关联（via 财务明细）──
    # 每个 (course, date) 只取第一个教练（正常情况一节课只有一个教练）
    session_coach_map = (
        finance_df[["_course", "_class_date", "_coach"]]
        .drop_duplicates(subset=["_course", "_class_date"])
        .set_index(["_course", "_class_date"])["_coach"]
        .to_dict()
    )
    class_df["_coach"] = class_df.apply(
        lambda r: session_coach_map.get((r["_course"], r["_class_date"])), axis=1
    )

    # 为超时明细补充教练信息，并按教练→日期排序
    for d in overtime_details:
        d["coach"] = str(session_coach_map.get((d["course"], d["date"]), "—"))
    overtime_details.sort(key=lambda d: (d.get("coach", ""), d.get("date", "")))

    # ── 5. 按教练计算到课率（应到人数 × 上课次数 / 已到学员数）──
    # 读取每个课程的应到人数配置
    class_capacity: Dict[str, int] = {}
    for course_k, cap_v in cfg.get("class_capacity", {}).items():
        cn = str(course_k).strip()
        try:
            v = int(cap_v)
            if cn and v > 0:
                class_capacity[cn] = v
        except (TypeError, ValueError):
            pass

    # 排除新开班 / 关键词 / 放假日期后的上课记录（含教练信息）
    non_excl = class_df[
        ~class_df.apply(
            lambda r: _is_excluded_from_attendance(r["_course"], r["_class_date"]), axis=1
        ) & class_df["_coach"].notna()
    ].copy()

    # 记录未配置应到人数的课程（供前端警告）
    all_courses_in_records: set = set(non_excl["_course"].unique())
    missing_capacity_courses: List[str] = sorted(all_courses_in_records - set(class_capacity.keys()))

    # 仅对已配置应到人数的课程计算到课率
    if class_capacity:
        non_excl_cfg = non_excl[non_excl["_course"].isin(class_capacity)].copy()
    else:
        non_excl_cfg = pd.DataFrame(columns=non_excl.columns)

    coach_attendance: Dict[str, Dict] = {}
    if len(non_excl_cfg) > 0:
        # 应到 = Σ (教练该课程唯一上课次数 × 课程应到人数)，按教练聚合
        uniq_sessions = (
            non_excl_cfg
            .drop_duplicates(subset=["_coach", "_course", "_class_date"])
            [["_coach", "_course"]]
            .copy()
        )
        uniq_sessions["_cap"] = uniq_sessions["_course"].map(class_capacity).fillna(0).astype(int)
        expected_by_coach = uniq_sessions.groupby("_coach")["_cap"].sum()

        # 已到 = 状态为「已到」的学员行数（仅已配置课程）
        attended_by_coach = (
            non_excl_cfg[non_excl_cfg["_attended"]]
            .groupby("_coach")
            .size()
        )

        for ck in set(expected_by_coach.index) | set(attended_by_coach.index):
            exp = int(expected_by_coach.get(ck, 0))
            att = int(attended_by_coach.get(ck, 0))
            coach_attendance[str(ck)] = {
                "rate":     att / exp if exp > 0 else 0.0,
                "attended": att,
                "total":    exp,
            }

    # ── 5b. 构建每位教练的到课率核算明细（供 Excel 导出）──
    attendance_sheets: Dict[str, Any] = {}
    has_member_col = "_member" in non_excl_cfg.columns if len(non_excl_cfg) > 0 else False
    if len(non_excl_cfg) > 0:
        for ck, ck_df in non_excl_cfg.groupby("_coach"):
            ck = str(ck)
            # 汇总：每门课唯一上课次数 × 应到人数
            sessions_per_course = (
                ck_df.drop_duplicates(subset=["_course", "_class_date"])
                .groupby("_course")
                .size()
            )
            course_summary: List[Dict] = []
            for course_name, n_sess in sessions_per_course.items():
                cap = class_capacity.get(str(course_name), 0)
                att_cnt = int(
                    ck_df[(ck_df["_course"] == course_name) & ck_df["_attended"]].shape[0]
                )
                course_summary.append({
                    "course":   str(course_name),
                    "sessions": int(n_sess),
                    "capacity": cap,
                    "subtotal": int(n_sess) * cap,
                    "attended": att_cnt,
                })
            course_summary.sort(key=lambda x: x["course"])

            # 明细记录（排序：课程→日期→学员）
            sort_cols = ["_course", "_class_date"]
            if has_member_col:
                sort_cols.append("_member")
            rec_rows: List[Dict] = []
            for _, r in ck_df.sort_values(sort_cols).iterrows():
                rec: Dict[str, str] = {
                    "course": str(r["_course"]),
                    "date":   str(r["_class_date"]),
                    "status": str(r["_status"]),
                }
                if has_member_col:
                    rec["student"] = str(r.get("_member", ""))
                rec_rows.append(rec)

            attendance_sheets[ck] = {
                "summary": course_summary,
                "records": rec_rows,
                "has_student": has_member_col,
            }

    # ── 6. 根据到课率确定阶梯单价 ──
    sorted_tiers = sorted(attendance_tiers, key=lambda t: float(t.get("min", 0)), reverse=True)

    def _get_normal_rate(coach: str, sport: str) -> float:
        info = coach_attendance.get(coach)
        rate = info["rate"] if info else None
        # 没有到课率数据（该教练无上课记录）→ 最低档
        best = sorted_tiers[-1] if sorted_tiers else {}
        if rate is not None:
            for t in sorted_tiers:
                if rate >= float(t.get("min", 0)):
                    best = t
                    break
        return float(best.get(sport, 0))

    # ── 7. 为财务明细行识别球类 + 课包类型 ──
    def _get_sport(row: pd.Series) -> str:
        pack = row.get("_pack_name", "")
        if pack and pack in course_pack_mapping:
            return course_pack_mapping[pack].get("sport", "unknown")
        return detect_sport(row["_course"], sport_keywords)

    def _get_pkg(row: pd.Series) -> str:
        pack = row.get("_pack_name", "")
        if pack and pack in course_pack_mapping:
            return course_pack_mapping[pack].get("pkg_type", "正常课包")
        return detect_package_type(row["_course"], row["_sport"], package_keywords)

    finance_df["_sport"]    = finance_df.apply(_get_sport, axis=1)
    finance_df["_pkg_type"] = finance_df.apply(_get_pkg, axis=1)

    # 超时标记：优先按学员精确匹配，缺会员列时退回整课匹配
    if overtime_student_set is not None and "_member" in finance_df.columns:
        finance_df["_is_overtime"] = finance_df.apply(
            lambda r: (r["_course"], r["_class_date"], r["_member"]) in overtime_student_set,
            axis=1,
        )
    else:
        finance_df["_is_overtime"] = finance_df.apply(
            lambda r: (r["_course"], r["_class_date"]) in overtime_set, axis=1
        )
    finance_df["_is_new_class"] = finance_df.apply(
        lambda r: _is_excluded_from_attendance(r["_course"], r["_class_date"]), axis=1
    )

    # ── 8. 计算每条财务明细的课时费 ──
    def _row_fee(row: pd.Series) -> float:
        if row["_is_overtime"]:
            return 0.0
        sport = row["_sport"]
        pkg   = row["_pkg_type"]
        coach = row["_coach"]
        if "168" in pkg:
            return 0.0
        if "私教" in pkg or "一对一" in pkg:
            # 费用 = unit_price × 费率；全部课次加总 = 总确认收入 × 费率
            return float(row["_unit_price"]) * one_on_one_rate
        if "跃动" in pkg:
            return yuedong_fee
        # 正常课包 & 新开班 → 同一阶梯
        if sport in ("football", "basketball"):
            return _get_normal_rate(coach, sport)
        return 0.0

    finance_df["_fee"] = finance_df.apply(_row_fee, axis=1)

    # ── 9. 获取每位教练的校区 + 师资（取众数）──
    coach_meta: Dict[str, Dict] = {}
    for coach, grp in finance_df.groupby("_coach"):
        campus_vals  = grp["_campus"][grp["_campus"].str.strip() != ""]
        quality_vals = grp["_quality"][grp["_quality"].str.strip() != ""]
        coach_meta[str(coach)] = {
            "campus":  campus_vals.mode().iloc[0]  if len(campus_vals)  > 0 else "",
            "quality": quality_vals.mode().iloc[0] if len(quality_vals) > 0 else "",
        }

    # ── 10. 按教练聚合 → coach_results ──
    def _sort_line(line: Dict) -> tuple:
        sport_order = {"football": 0, "basketball": 1, "unknown": 2}
        pkg_order   = {"正常课包": 0, "跃动课包": 1, "168拼团课包": 2, "168课包": 2, "一对一私教": 3}
        return (sport_order.get(line["sport"], 9), line["is_overtime"], pkg_order.get(line["pkg_type"], 9))

    coach_results: List[Dict] = []
    for coach, cgrp in finance_df.groupby("_coach"):
        coach = str(coach)
        meta  = coach_meta.get(coach, {})
        att   = coach_attendance.get(coach)

        lines: List[Dict] = []
        for (sport, pkg, is_ot), lgrp in cgrp.groupby(["_sport", "_pkg_type", "_is_overtime"]):
            count     = len(lgrp)
            total_fee = round(float(lgrp["_fee"].sum()), 2)
            unit_fee  = round(total_fee / count, 4) if count > 0 else 0.0

            sport_label = SPORT_LABELS.get(sport, sport if sport != "unknown" else "未识别")
            ot_suffix   = "-超时-无课时费" if is_ot else ""
            label       = f"{sport_label}-{pkg}{ot_suffix}"

            lines.append({
                "label":      label,
                "sport":      sport,
                "pkg_type":   pkg,
                "is_overtime": bool(is_ot),
                "count":      count,
                "unit_fee":   unit_fee,
                "total_fee":  total_fee,
            })

        lines.sort(key=_sort_line)
        total = round(sum(l["total_fee"] for l in lines), 2)

        coach_results.append({
            "coach":            coach,
            "campus":           meta.get("campus", ""),
            "quality":          meta.get("quality", ""),
            "attendance_rate":  att["rate"]     if att else None,
            "attendance_stats": {"attended": att["attended"], "total": att["total"]} if att else None,
            "lines":            lines,
            "total":            total,
        })

    coach_results.sort(key=lambda r: r["coach"])

    # ── 11. 统计 ──
    grand_total = round(sum(r["total"] for r in coach_results), 2)
    stats = {
        "coach_count":           len(coach_results),
        "total_fee":             grand_total,
        "session_count":         len(finance_df),
        "overtime_count":        int(finance_df["_is_overtime"].sum()),
        "unmatched_sport_courses": sorted(
            finance_df[finance_df["_sport"] == "unknown"]["_course"].unique().tolist()
        ),
        "missing_capacity_courses": missing_capacity_courses,
        "overtime_details":          overtime_details,
        "overtime_threshold_days":   threshold,
        "overtime_student_count":    sum(d["late_count"] for d in overtime_details),
        # 超时判定粒度：per_student = 按学员精确匹配，per_session = 整课匹配（兼容/无会员列）
        "overtime_match_mode": "per_student" if overtime_student_set is not None else "per_session",
        "overtime_match_debug": {
            "class_member_col":   col_cls_member,
            "finance_member_col": col_f_member,
        },
    }

    return {
        "coach_results":    coach_results,
        "stats":            stats,
        "attendance_sheets": attendance_sheets,
    }


# ──────────────────── Excel 导出 ────────────────────

def export_to_excel(
    coach_results: List[dict],
    stats: dict,
    attendance_sheets: Optional[Dict[str, Any]] = None,
) -> bytes:
    """生成 Excel：
    - Sheet1「课时费明细」：汇总表
    - 每位教练一个 Sheet：到课率核算明细（课程汇总 + 逐条上课记录）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "课时费明细"

    # ── 样式定义 ──
    header_fill  = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
    total_fill   = PatternFill(start_color="D9E8D5", end_color="D9E8D5", fill_type="solid")
    ot_fill      = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    grand_fill   = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    bold_font    = Font(bold=True, size=10)
    grand_font   = Font(color="FFFFFF", bold=True, size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS = ["序号", "校区", "姓名", "师资", "到课率", "计费课包", "应付课时", "金额", "个人课时合计", "应发课时费"]
    ws.append(COLS)

    # 设置表头样式
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center

    # 列宽
    col_widths = [6, 20, 10, 8, 10, 24, 10, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 2
    for idx, coach in enumerate(coach_results, 1):
        lines = coach["lines"]
        span  = len(lines)
        rate  = coach["attendance_rate"]
        rate_str = f"{rate:.2%}" if rate is not None else "-"

        for li, line in enumerate(lines):
            row_data = [
                idx           if li == 0 else "",
                coach["campus"]   if li == 0 else "",
                coach["coach"]    if li == 0 else "",
                coach["quality"]  if li == 0 else "",
                rate_str      if li == 0 else "",
                line["label"],
                line["count"],
                f"{line['unit_fee']:.2f}",
                f"{coach['total']:.2f}" if li == 0 else "",
                f"{coach['total']:.2f}" if li == 0 else "",
            ]
            ws.append(row_data)
            r = ws[current_row]

            # 超时行 → 黄色底色
            if line["is_overtime"]:
                for cell in r:
                    cell.fill = ot_fill

            for i, cell in enumerate(r):
                cell.alignment = center if i not in (1, 2, 5) else left

            current_row += 1

        # 合并教练信息列（序号、校区、姓名、师资、到课率、个人合计、应发）
        if span > 1:
            merge_cols = [1, 2, 3, 4, 5, 9, 10]
            for col in merge_cols:
                ws.merge_cells(
                    start_row=current_row - span, end_row=current_row - 1,
                    start_column=col,             end_column=col,
                )
            # 合并后重新设置合并单元格对齐
            for col in merge_cols:
                ws.cell(row=current_row - span, column=col).alignment = center

        # 教练小计行的字体加粗
        for col in [9, 10]:
            ws.cell(row=current_row - span, column=col).font = bold_font

    # ── 合计行 ──
    grand_total = stats.get("total_fee", 0)
    ws.append(["合计", "", "", "", "", "", stats.get("session_count", ""), "",
               f"{grand_total:.2f}", f"{grand_total:.2f}"])
    for cell in ws[current_row]:
        cell.fill      = grand_fill
        cell.font      = grand_font
        cell.alignment = center
    ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=5)

    # 冻结首行
    ws.freeze_panes = "A2"

    # ── 每位教练的到课率核算 Sheet ──
    if attendance_sheets:
        attend_fill   = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        section_font  = Font(bold=True, size=10, color="4A7C59")
        title_font    = Font(bold=True, size=12)

        for coach_name, sheet_data in attendance_sheets.items():
            # Sheet 名最长 31 字符
            safe_name = (coach_name[:29] + "…") if len(coach_name) > 31 else coach_name
            ws_c = wb.create_sheet(title=safe_name)

            summary     = sheet_data.get("summary", [])
            records     = sheet_data.get("records", [])
            has_student = sheet_data.get("has_student", False)

            # 标题行
            ws_c.append([f"{coach_name} — 到课率核算"])
            title_cell = ws_c.cell(row=1, column=1)
            title_cell.font      = title_font
            title_cell.alignment = left
            ws_c.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
            ws_c.append([])

            # ── 汇总表 ──
            ws_c.append(["【应到人数汇总】"])
            ws_c.cell(row=ws_c.max_row, column=1).font = section_font

            sum_header = ["课程名称", "上课次数", "应到人数/次", "应到小计（次×人）", "实到人数"]
            ws_c.append(sum_header)
            for cell in ws_c[ws_c.max_row]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center

            total_subtotal = 0
            total_attended = 0
            for row in summary:
                ws_c.append([
                    row["course"],
                    row["sessions"],
                    row["capacity"],
                    row["subtotal"],
                    row["attended"],
                ])
                r = ws_c.max_row
                ws_c.cell(r, 1).alignment = left
                for col in range(2, 6):
                    ws_c.cell(r, col).alignment = center
                total_subtotal += row["subtotal"]
                total_attended += row["attended"]

            # 合计行
            rate_val = total_attended / total_subtotal if total_subtotal > 0 else 0.0
            ws_c.append(["合计", "—", "—",
                          total_subtotal, total_attended,
                          f"{rate_val:.2%}"])
            r_total = ws_c.max_row
            for cell in ws_c[r_total]:
                cell.fill      = total_fill
                cell.font      = bold_font
                cell.alignment = center

            ws_c.append([])

            # ── 明细记录 ──
            ws_c.append(["【上课记录明细】"])
            ws_c.cell(row=ws_c.max_row, column=1).font = section_font

            rec_header = ["课程名称", "上课日期"]
            if has_student:
                rec_header.append("学员姓名")
            rec_header.append("状态")
            ws_c.append(rec_header)
            for cell in ws_c[ws_c.max_row]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center

            for rec in records:
                row_data = [rec["course"], rec["date"]]
                if has_student:
                    row_data.append(rec.get("student", ""))
                row_data.append(rec["status"])
                ws_c.append(row_data)
                r = ws_c.max_row
                ws_c.cell(r, 1).alignment = left
                status_col = 4 if has_student else 3
                if rec["status"] == "已到":
                    ws_c.cell(r, status_col).fill = attend_fill
                ws_c.cell(r, status_col).alignment = center
                if len(row_data) > 1:
                    ws_c.cell(r, 2).alignment = center

            # 列宽
            ws_c.column_dimensions["A"].width = 36
            ws_c.column_dimensions["B"].width = 14
            ws_c.column_dimensions["C"].width = 14
            ws_c.column_dimensions["D"].width = 16
            ws_c.column_dimensions["E"].width = 12
            ws_c.column_dimensions["F"].width = 10
            ws_c.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
