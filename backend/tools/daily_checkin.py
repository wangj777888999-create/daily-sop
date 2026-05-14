import io
import json
import os
import logging
from typing import Dict, List, Tuple
from datetime import date

import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data", "config.json")


def _load_config() -> dict:
    if not os.path.exists(CONFIG_PATH):
        raise RuntimeError(f"配置文件缺失，请确保 data/config.json 存在: {CONFIG_PATH}")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# Chinese → English column mapping for direct import
COLUMN_MAP = {
    "部门": "department",
    "学校名称": "school_name",
    "课程类型": "course_type",
    "课程名称": "course_name",
    "教练姓名": "coach_name",
    "课程日期": "course_date",
    "开课时间": "start_time",
    "下课时间": "end_time",
    "签到时间": "sign_in_time",
    "签退时间": "sign_out_time",
    "签到状态": "sign_status",
    "实际上课人次": "actual_count",
    "课程应到人次": "expected_count",
    "确认收入": "confirmed_revenue",
}


def import_checkin(excel_bytes: bytes, check_date: str = None) -> dict:
    """Import a single pre-formatted Excel directly into records.

    The Excel should have Chinese column headers matching COLUMN_MAP.
    Accepts both raw coach-checkin format (no finance columns) and
    fully-processed output format (with finance columns).
    """
    import pandas as pd
    from tools.course_types import load_all as ct_load_all

    xls = pd.ExcelFile(io.BytesIO(excel_bytes))
    sheet_name = xls.sheet_names[0]
    raw = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=sheet_name, header=None)

    # Auto-detect header row: find the first row that contains at least 3 known column names
    header_row_idx = None
    for row_idx in range(min(8, len(raw))):
        row_vals = set(str(v) for v in raw.iloc[row_idx].values if isinstance(v, str))
        hits = sum(1 for k in COLUMN_MAP if k in row_vals)
        if hits >= 3:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        available = ", ".join(str(v) for v in raw.iloc[0].values if isinstance(v, str))
        raise ValueError(
            f"无法识别列名，文件前几行未找到预期的中文列名。"
            f"第一行内容: {available}。预期列名: {', '.join(COLUMN_MAP.keys())}"
        )

    # Skip rows above header, use header row as column names
    raw.columns = raw.iloc[header_row_idx].values
    df = raw.iloc[header_row_idx + 1:].reset_index(drop=True)

    # Filter to school courses only (skip aggregate/other rows)
    if "课程类型" in df.columns:
        df = df[df["课程类型"] == "学校课程"]

    if df.empty:
        raise ValueError("文件中未找到「学校课程」类型的签到记录，请检查文件内容")

    # Map columns and build records
    records = []
    for _, row in df.iterrows():
        record = {
            "check_date": check_date,
            "department": "",
            "school_name": "",
            "course_type": "",
            "course_name": "",
            "coach_name": "",
            "course_date": "",
            "start_time": "",
            "end_time": "",
            "sign_in_time": "",
            "sign_out_time": "",
            "sign_status": "",
            "actual_count": 0,
            "expected_count": 0,
            "confirmed_revenue": 0.0,
        }
        for cn, en in COLUMN_MAP.items():
            if cn in df.columns:
                val = row.get(cn)
                if pd.isna(val):
                    continue
                if en in ("actual_count", "expected_count"):
                    record[en] = int(float(val))
                elif en == "confirmed_revenue":
                    record[en] = float(val)
                else:
                    record[en] = str(val)

        # Each record uses its own course_date as check_date; fallback to user-provided date
        if record["course_date"]:
            record["check_date"] = record["course_date"]
        elif not record["check_date"]:
            record["check_date"] = check_date

        records.append(record)

    if not check_date:
        check_date = str(date.today())

    # Generate output Excel from the records (normalized format)
    cols_order = [
        "部门", "学校名称", "课程类型", "课程名称", "教练姓名",
        "实际上课人次", "课程应到人次", "确认收入",
        "课程日期", "开课时间", "下课时间", "签到时间", "签退时间", "签到状态",
    ]
    en_to_cn = {v: k for k, v in COLUMN_MAP.items()}
    out_data = []
    for r in records:
        out_row = {}
        for cn in cols_order:
            en = COLUMN_MAP.get(cn)
            out_row[cn] = r.get(en, "") if en else ""
        out_data.append(out_row)

    out_df = pd.DataFrame(out_data, columns=cols_order)
    excel_bytes_out = _generate_excel(out_df)

    # Summary
    dept_set = set(r["department"] for r in records if r["department"])
    school_set = set(r["school_name"] for r in records if r["school_name"])
    coach_set = set(r["coach_name"] for r in records if r["coach_name"])
    non_gang = sum(1 for r in records if r["sign_status"] != "在岗")

    # Check unmapped course types
    unmapped_courses = []
    try:
        all_courses = set(r["course_name"] for r in records if r["course_name"])
        ct_records = ct_load_all()
        mapped = {r["课程名称"] for r in ct_records}
        unmapped_courses = sorted(all_courses - mapped)
    except Exception:
        pass

    return {
        "records": records,
        "excel_bytes": excel_bytes_out,
        "summary": {
            "total_records": len(records),
            "departments": len(dept_set),
            "schools": len(school_set),
            "coaches": len(coach_set),
            "non_岗_count": non_gang,
        },
        "unmapped_courses": unmapped_courses,
    }


def preview_files(coach_bytes: bytes, finance_bytes: bytes) -> dict:
    """Parse uploaded Excel bytes and return preview (first 50 rows + column validation)."""
    coach_df = pd.read_excel(io.BytesIO(coach_bytes), sheet_name="教练签到")
    finance_df = pd.read_excel(io.BytesIO(finance_bytes), sheet_name="财务")

    # Skip header rows like the original script
    coach_data = coach_df.iloc[2:].copy()
    coach_data.columns = coach_df.iloc[1].values

    finance_data = finance_df.iloc[3:].copy()
    finance_data.columns = [
        '学员类型', '学员姓名', '订单类型', '课包名称', '订单金额', '优惠金额',
        '付款金额', '课包课次', '课包赠送课次', '课包已上总课次', '已确认总收入',
        '退款金额', '退款确认收入金额', '课程名称', '课程明细_换班前课程',
        '课程明细_兑换金额', '课程明细_课程单价', '课程明细_总课次(缴费总课次)',
        '课程明细_赠送课次', '课程明细_已上总课次', '课程明细_已确认总收入',
        '课程明细_剩余总课次', '课程明细_总课次(选定时间)', '课程明细_已上课次（选定时间）',
        '课程明细_已确认收入（选定时间）', '课程明细_剩余课次（选定时间）', '课程明细_已确认收入的详情'
    ]

    return {
        "coach_columns": list(coach_data.columns),
        "coach_rows": coach_data.head(50).fillna("").to_dict(orient="records"),
        "coach_total": len(coach_data),
        "finance_columns": list(finance_data.columns),
        "finance_rows": finance_data.head(50).fillna("").to_dict(orient="records"),
        "finance_total": len(finance_data),
    }


def process_files(coach_bytes: bytes, finance_bytes: bytes, check_date: str = None) -> dict:
    """Process coach check-in + finance data. Returns {records, excel_bytes, summary}."""
    config = _load_config()
    dept_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        config["tools"]["department_mapping"]
    )

    # Parse coach data
    coach_df = pd.read_excel(io.BytesIO(coach_bytes), sheet_name="教练签到")
    coach_data = coach_df.iloc[2:].copy()
    coach_data.columns = coach_df.iloc[1].values
    coach_data = coach_data.query("课程类型 == '学校课程'").drop(columns='场馆名称', errors='ignore')

    # Parse finance data
    finance_df = pd.read_excel(io.BytesIO(finance_bytes), sheet_name="财务")
    finance_data = finance_df.iloc[3:].copy()
    finance_data.columns = [
        '学员类型', '学员姓名', '订单类型', '课包名称', '订单金额', '优惠金额',
        '付款金额', '课包课次', '课包赠送课次', '课包已上总课次', '已确认总收入',
        '退款金额', '退款确认收入金额', '课程名称', '课程明细_换班前课程',
        '课程明细_兑换金额', '课程明细_课程单价', '课程明细_总课次(缴费总课次)',
        '课程明细_赠送课次', '课程明细_已上总课次', '课程明细_已确认总收入',
        '课程明细_剩余总课次', '课程明细_总课次(选定时间)', '课程明细_已上课次（选定时间）',
        '课程明细_已确认收入（选定时间）', '课程明细_剩余课次（选定时间）', '课程明细_已确认收入的详情'
    ]

    numeric_cols = ['课程明细_已确认收入（选定时间）', '课程明细_已上课次（选定时间）', '课程明细_总课次(选定时间)']
    for col in numeric_cols:
        finance_data[col] = pd.to_numeric(finance_data[col], errors='coerce').fillna(0)

    fi = finance_data.groupby('课程名称').agg(
        实际上课人次=('课程明细_已上课次（选定时间）', 'sum'),
        课程应到人次=('课程明细_总课次(选定时间)', 'sum'),
        确认收入=('课程明细_已确认收入（选定时间）', 'sum')
    ).reset_index()

    # Load department mapping
    dept_df = pd.read_excel(dept_path, sheet_name='Sheet1')
    department = dept_df.rename(columns={'学校': '学校名称'})

    # Merge all data
    merged = pd.merge(coach_data, department[['学校名称', '部门']], on='学校名称', how='left')
    result = pd.merge(merged, fi, on='课程名称', how='left')

    # 左连接后财务列可能为 NaN（课程在财务表中无匹配），统一填 0
    for col in ['实际上课人次', '课程应到人次', '确认收入']:
        if col in result.columns:
            result[col] = result[col].fillna(0)

    # Reorder columns
    new_order = ['部门']
    for col in result.columns:
        if col not in new_order and col not in ['实际上课人次', '课程应到人次', '确认收入']:
            new_order.append(col)
    if '教练姓名' in new_order:
        coach_index = new_order.index('教练姓名')
        new_order.insert(coach_index + 1, '实际上课人次')
        new_order.insert(coach_index + 2, '课程应到人次')
        new_order.insert(coach_index + 3, '确认收入')
    else:
        new_order.extend(['实际上课人次', '课程应到人次', '确认收入'])

    result = result[new_order].sort_values(by='部门', na_position='last')

    # Determine check_date
    if not check_date:
        check_date = str(date.today())

    # Convert to records for SQLite
    records = []
    for _, row in result.iterrows():
        records.append({
            "check_date": check_date,
            "department": str(row.get("部门", "")),
            "school_name": str(row.get("学校名称", "")),
            "course_type": str(row.get("课程类型", "")),
            "course_name": str(row.get("课程名称", "")),
            "coach_name": str(row.get("教练姓名", "")),
            "course_date": str(row.get("课程日期", "")),
            "start_time": str(row.get("开课时间", "")),
            "end_time": str(row.get("下课时间", "")),
            "sign_in_time": str(row.get("签到时间", "")),
            "sign_out_time": str(row.get("签退时间", "")),
            "sign_status": str(row.get("签到状态", "")),
            "actual_count": int(pd.to_numeric(row.get("实际上课人次", 0), errors='coerce') or 0),
            "expected_count": int(pd.to_numeric(row.get("课程应到人次", 0), errors='coerce') or 0),
            "confirmed_revenue": float(pd.to_numeric(row.get("确认收入", 0), errors='coerce') or 0),
        })

    # Generate formatted Excel
    excel_bytes = _generate_excel(result)

    # Summary
    status_col = None
    for col in ['状态', '签到状态', '教练状态']:
        if col in result.columns:
            status_col = col
            break

    summary = {
        "total_records": len(records),
        "departments": result['部门'].nunique() if '部门' in result.columns else 0,
        "schools": result['学校名称'].nunique() if '学校名称' in result.columns else 0,
        "coaches": result['教练姓名'].nunique() if '教练姓名' in result.columns else 0,
        "non_岗_count": int((result[status_col] != '在岗').sum()) if status_col and status_col in result.columns else 0,
    }

    # 检测未映射课程类型的课程
    unmapped_courses = []
    if '课程名称' in result.columns:
        all_courses = set(result['课程名称'].dropna().unique())
        try:
            from tools.course_types import load_all
            ct_records = load_all()
            mapped = {r['课程名称'] for r in ct_records}
            unmapped_courses = sorted(all_courses - mapped)
        except Exception:
            pass

    return {
        "records": records,
        "excel_bytes": excel_bytes,
        "summary": summary,
        "unmapped_courses": unmapped_courses,
    }


def _generate_excel(df: pd.DataFrame) -> bytes:
    """Generate formatted Excel with highlighting for non-在岗 statuses."""
    wb = Workbook()
    ws = wb.active
    ws.title = "cwcl"

    headers = list(df.columns)
    ws.append(headers)

    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    modified_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')

    status_col_name = None
    for col in ['状态', '签到状态', '教练状态']:
        if col in headers:
            status_col_name = col
            break
    status_col_idx = headers.index(status_col_name) + 1 if status_col_name else None

    # Columns where 0 means "no data" → display blank
    zero_as_blank_cols = {'实际上课人次', '课程应到人次', '确认收入'}

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        row_values = row.tolist()
        # Replace 0 with empty string for zero-as-blank columns
        for i, h in enumerate(headers):
            if h in zero_as_blank_cols and row_values[i] == 0:
                row_values[i] = ''
        ws.append(row_values)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
        if status_col_idx and row[status_col_name] != '在岗':
            ws.cell(row=row_idx, column=status_col_idx).fill = highlight_fill
        # Highlight status cell if it was manually modified
        if status_col_idx and '备注' in headers:
            remark = str(row.get('备注', '') or '')
            if '[状态已修改' in remark:
                ws.cell(row=row_idx, column=status_col_idx).fill = modified_fill

    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).alignment = center_alignment

    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 45

    # Auto column width
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell_value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_length + 2, 10), 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
