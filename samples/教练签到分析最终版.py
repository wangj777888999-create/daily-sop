import pandas as pd
from openpyxl.styles import PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def process_coach_data(coach):
    """处理教练签到数据"""
    # 跳过前两行，设置新列名
    new_header = coach.iloc[1].values
    coach_data = coach.iloc[2:].copy()
    coach_data.columns = new_header
    
    # 筛选并清理数据
    return coach_data.query("课程类型 == '学校课程'").drop(columns='场馆名称', errors='ignore')

def process_finance_data(finance):
    """处理财务统计数据"""
    # 跳过前三行
    finance = finance.iloc[3:].copy()
    
    # 设置列名
    finance.columns = [
        '学员类型', '学员姓名', '订单类型', '课包名称', '订单金额', '优惠金额', 
        '付款金额', '课包课次', '课包赠送课次', '课包已上总课次', '已确认总收入', 
        '退款金额', '退款确认收入金额', '课程名称', '课程明细_换班前课程', 
        '课程明细_兑换金额', '课程明细_课程单价', '课程明细_总课次(缴费总课次)', 
        '课程明细_赠送课次', '课程明细_已上总课次', '课程明细_已确认总收入', 
        '课程明细_剩余总课次', '课程明细_总课次(选定时间)', '课程明细_已上课次（选定时间）', 
        '课程明细_已确认收入（选定时间）', '课程明细_剩余课次（选定时间）', '课程明细_已确认收入的详情'
    ]
    # finance = finance[pd.to_numeric(finance['课程明细_已确认收入（选定时间）'], errors='coerce') != 0]
    numeric_cols = ['课程明细_已确认收入（选定时间）', '课程明细_已上课次（选定时间）', '课程明细_总课次(选定时间)']
    for col in numeric_cols:
        finance[col] = pd.to_numeric(finance[col], errors='coerce').fillna(0)
    
    # 转换数据类型并聚合
    columns = ['课程明细_已确认收入（选定时间）', '课程明细_已上课次（选定时间）', '课程明细_总课次(选定时间)']
    for col in columns:
        finance[col] = pd.to_numeric(finance[col], errors='coerce')
    
    fi = finance.groupby('课程名称').agg(
        实际上课人次=('课程明细_已上课次（选定时间）', 'sum'),
        课程应到人次=('课程明细_总课次(选定时间)', 'sum'),
        确认收入=('课程明细_已确认收入（选定时间）', 'sum')
    ).reset_index()
    print(print(fi['课程名称'].unique()))
    return fi
def merge_and_reorder_data(coach_data, coach_department, finance_data):
    """合并所有数据源并调整列顺序"""
    # 重命名部门数据列
    department = coach_department.rename(columns={'学校': '学校名称'})
    
    # 合并教练数据和部门数据
    merged = pd.merge(
        coach_data, 
        department[['学校名称', '部门']], 
        on='学校名称', 
        how='left'
    )
    
    # 合并财务数据
    result = pd.merge(
        merged, 
        finance_data, 
        on='课程名称', 
        how='left'
    )
    
    # 定义新的列顺序
    new_order = ['部门']
    
    # 添加其他列（除了已处理的列）
    for col in result.columns:
        if col not in new_order and col not in ['实际上课人次', '课程应到人次', '确认收入']:
            new_order.append(col)
    
    # 在教练姓名后插入三个指标列
    if '教练姓名' in new_order:
        coach_index = new_order.index('教练姓名')
        new_order.insert(coach_index + 1, '实际上课人次')
        new_order.insert(coach_index + 2, '课程应到人次')
        new_order.insert(coach_index + 3, '确认收入')
    else:
        # 如果找不到教练姓名，则将三个指标列放在最后
        new_order.extend(['实际上课人次', '课程应到人次', '确认收入'])
    
    return result[new_order]

def calculate_column_width(ws):
    """计算每列的最佳宽度"""
    column_widths = {}
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row, col).value
            if cell_value:
                # 计算字符串长度（中文字符算2个宽度）
                length = sum(2 if ord(char) > 127 else 1 for char in str(cell_value))
                if length > max_length:
                    max_length = length
        # 设置最小宽度为10，最大宽度为50，并添加缓冲
        column_widths[col] = min(max(max_length + 2, 10), 50)
    return column_widths

def export_with_formatting(df, output_file):
    """导出数据并应用格式和高亮显示"""
    wb = Workbook()
    ws = wb.active
    ws.title = "cwcl"
    
    # 写入列标题
    headers = list(df.columns)
    ws.append(headers)
    
    # 设置样式
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 查找状态列
    status_col_name = next((col for col in ['状态', '签到状态', '教练状态'] if col in headers), None)
    status_col_idx = headers.index(status_col_name) + 1 if status_col_name else None
    
    # 写入数据
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        ws.append(row.tolist())
        
        # 应用居中对齐
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
            
        # 高亮状态不为"在岗"的单元格
        if status_col_idx and row[status_col_name] != '在岗':
            ws.cell(row=row_idx, column=status_col_idx).fill = highlight_fill
    
    # 设置标题行格式
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = center_alignment
    
    # 设置行高
    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 45
    
    # 计算并设置自适应列宽
    column_widths = calculate_column_width(ws)
    for col, width in column_widths.items():
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = width
    
    wb.save(output_file)

def main():
    """主函数"""
    # 读取数据
    coach_department = pd.read_excel('部门划分.xlsx', sheet_name='Sheet1')
    coach = pd.read_excel(r'C:\Users\Administrator\Desktop\python\2026年4月签到数据\4.21\教练签到 (53).xls', sheet_name='教练签到')
    finance = pd.read_excel(r'C:\Users\Administrator\Desktop\python\2026年4月签到数据\4.21\财务统计 (78).xls', sheet_name='财务')
    
    # 处理数据
    coach_data = process_coach_data(coach)
    finance_data = process_finance_data(finance)
    result = merge_and_reorder_data(coach_data, coach_department, finance_data)
    # 根据部门排序
    result = result.sort_values(by='部门', na_position='last')
    
    # 导出并应用格式
    export_with_formatting(result, r'C:\Users\Administrator\Desktop\python\2026年4月签到数据\4月教练签到\4.21.xlsx')
    print("数据处理完成，结果已保存到'10.24.xlsx'")

if __name__ == "__main__":
    main()