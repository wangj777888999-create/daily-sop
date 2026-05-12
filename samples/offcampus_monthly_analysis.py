import pandas as pd
import numpy as np
from collections import defaultdict
import os
import datetime
from dateutil.relativedelta import relativedelta  # 需安装：pip install python-dateutil
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
# from tabulate import tabulate
def skjl_sperate(skjl,cw):
    outside_kw = ['农都', '西溪', '滨江']
    skjl['是否校外'] = skjl['课程名称'].str.contains('|'.join(outside_kw), case=False, na=False).copy()
    xwskjl = skjl[skjl['课程名称'].str.contains('|'.join(outside_kw), case=False, na=False)].copy()
    xwskjl['时加课'] = xwskjl['上课信息'].astype(str) + xwskjl['课程名称'].astype(str)
    cw_copy ,curo= process_finance_data(cw)
    # print(xwskjl.info())
    cw_copy['时加课'] = cw_copy['上课日期'].astype(str) + ' ' + cw_copy['上课时间'].astype(str) + cw_copy['课程名称'].astype(str)
    cw_copy = cw_copy.drop_duplicates(subset=['时加课'])[['时加课','校区']].copy()
    xwskjl = pd.merge(xwskjl, cw_copy[['时加课','校区']], on='时加课', how='left')
    # print(xwskjl.info())
    xwskjl = pd.merge(xwskjl, curo[['时加课','教练']], on='时加课', how='left')
    xwskjl.rename(columns={'订单类型':'校区'}, inplace=True)
    xwskjl_copy = xwskjl.drop(columns=['时加课']).copy()
    xnskjl = skjl[~skjl['课程名称'].str.contains('|'.join(outside_kw), case=False, na=False)].copy()
    return xwskjl_copy, xnskjl
# --------------------------
# 子函数：1. 生成教练归属表（区分校外/校内）
# --------------------------
def create_coach_belonging(cw, off_campus_coaches):
    """生成教练归属表（1=校外，0=校内）"""
    all_coaches = cw['教练'].drop_duplicates().tolist()
    return pd.DataFrame({
        '教练': all_coaches,
        '所属': [1 if coach in off_campus_coaches else 0 for coach in all_coaches]
    })


# --------------------------
# 子函数：2. 处理财务数据并生成课程-教练映射
# --------------------------
def process_finance_data(cw):
    """处理财务数据，生成校外财务表和课程-教练映射关系"""
    # 校外财务数据（球馆学员）
    xwcw = cw[cw['学员类型'] == '球馆学员'].copy()
    xwcw.rename(columns={'订单类型': '校区'}, inplace=True)
    
    # 生成课程唯一标识（时加课）及课程-教练映射
    cw_copy = cw.copy()
    cw_copy['时加课'] = cw_copy['上课日期'].astype(str) + ' ' + cw_copy['上课时间'].astype(str) + cw_copy['课程名称'].astype(str)
    course_coach = cw_copy.drop_duplicates(subset=['时加课'])[['时加课', '教练']].copy()
    # course_campus = cw_copy.drop_duplicates(subset=['时加课'])[['时加课','校区']].copy()
    return xwcw, course_coach


# --------------------------
# 子函数：3. 计算基础统计（课次、人次、平均课堂人次）
# --------------------------
def calculate_basic_stats(xwskjl):
    """计算上课节次、人次、平均课堂人次"""
    # print(xwskjl.info())
    xwskjl_copy = xwskjl.copy()
    # 生成课程唯一标识和校区教练组合键
    xwskjl_copy['时加课'] = xwskjl_copy['上课信息'].astype(str) + xwskjl_copy['课程名称'].astype(str)
    xwskjl_copy['校区教练'] = xwskjl_copy['校区'].astype(str) + '-' + xwskjl_copy['教练'].astype(str)
    
    # 筛选已到记录
    xwskjl_in = xwskjl_copy[xwskjl_copy['状态'] == '已到'].copy()
    xwskjl_in['校区教练'] = xwskjl_in['校区'].astype(str) + '-' + xwskjl_in['教练'].astype(str)
    
    # 统计上课节次（去重课程）
    campus_course = xwskjl_in.drop_duplicates(subset=['时加课']).groupby(
        '校区教练', as_index=False
    )['时加课'].nunique().rename(columns={'时加课': '上课节次'})
    
    # 统计上课人次
    campus_people = xwskjl_in.groupby('校区教练', as_index=False).size().rename(columns={'size': '上课人次'})
    
    # 合并并计算平均课堂人次
    combined = pd.merge(campus_course, campus_people, on='校区教练', how='outer')
    combined['上课节次'] = combined['上课节次'].fillna(0).astype(int)
    combined['上课人次'] = combined['上课人次'].fillna(0).astype(int)
    combined[['校区', '教练']] = combined['校区教练'].str.split('-', expand=True)
    combined['平均课堂人次'] = combined.apply(
        lambda x: round(x['上课人次']/x['上课节次'], 2) if x['上课节次'] != 0 else 0, axis=1
    )
    
    return combined, xwskjl_copy  # 返回基础统计结果和处理后的xwskjl


# --------------------------
# 子函数：4. 计算到课率
# --------------------------
def calculate_attendance_rate(combined, xwskjl_copy):
    """计算到课率并合并到主表"""
    # 总应到人数
    xwskjl_filtered = xwskjl_copy[~xwskjl_copy['课程名称'].str.contains('私教', case=False, na=False)].copy()
    xwskjl_in  = xwskjl_filtered[xwskjl_filtered['状态'] == '已到'].copy()
    total_persons = xwskjl_filtered.groupby('校区教练', as_index=False).size().rename(columns={'size': '总应到人数'})
    # 已到人数（复用基础统计中的上课人次）
    attended_persons = xwskjl_in.groupby('校区教练', as_index=False).size().rename(columns={'size': '已到人数'})
    # 计算到课率
    attendance = pd.merge(total_persons, attended_persons, on='校区教练', how='left')
    # print(attendance.info())
    attendance['已到人数'] = attendance['已到人数'].fillna(0)
    attendance['临时到课率'] = attendance.apply(
        lambda x: round(x['已到人数']/x['总应到人数'], 4) if x['总应到人数'] != 0 else 0, axis=1
    )
    # valid_mask = ~((attendance['已到人数'] == 1) & (attendance['临时到课率'] == 1.0))
    # attendance_valid = attendance[valid_mask].copy()
    
    # print(attendance_valid.info())
    attendance['到课率'] = attendance['临时到课率'].apply(lambda x: f"{x*100:.2f}%")
   
    combined = pd.merge(combined, attendance[['校区教练', '到课率']], on='校区教练', how='left')
    print(combined)
    # 合并到主表
    return combined


# --------------------------
# 子函数：5. 统计确认收入与平均课程单价
# --------------------------
def calculate_revenue(combined, xwcw):
    """计算确认收入和平均课程单价"""
    xwcw_copy = xwcw.copy()
    xwcw_copy['校区教练'] = xwcw_copy['校区'].astype(str) + '-' + xwcw_copy['教练'].astype(str)
    
    # 统计确认收入（课程单价总和）
    campus_revenue = xwcw_copy.groupby('校区教练', as_index=False)['课程单价'].sum().rename(columns={'课程单价': '确认收入'})
    campus_revenue['确认收入'] = campus_revenue['确认收入'].fillna(0).round(2)
    
    # 合并收入数据并计算平均课程单价
    combined = pd.merge(combined, campus_revenue[['校区教练', '确认收入']], on='校区教练', how='left')
    combined['确认收入'] = combined['确认收入'].fillna(0).round(2)
    combined['平均课程单价'] = combined.apply(
        lambda x: round(x['确认收入']/x['上课人次'], 2) if x['上课人次'] != 0 else 0, axis=1
    )
    combined.rename(columns={'上课节次': '课次'}, inplace=True)  # 统一列名
    
    return combined


# --------------------------
# 子函数：6. 分析五人以下/十人以上课次
# --------------------------
def analyze_class_size(combined, xwskjl_copy):
    """统计5人及以下、10人及以上课程数及比例"""
    # 按课程和校区教练分组，统计已到人数
    xwskjl_filtered = xwskjl_copy[~xwskjl_copy['课程名称'].str.contains('私教', case=False, na=False)].copy()

    ca_grouped = xwskjl_filtered.groupby(['上课信息', '校区教练'])['状态'].apply(list).reset_index()
    # print(ca_grouped.info())
    ca_container = defaultdict(lambda: {'total': 0, 'low': 0, 'high': 0})  # 用defaultdict减少重复初始化
    # print(ca_grouped.info())
    for _, row in ca_grouped.iterrows():
        present = sum(1 for s in row['状态'] if s == '已到')
        ca_container[row['校区教练']]['total'] += 1
        if present <= 5:
            ca_container[row['校区教练']]['low'] += 1
        if present >= 10:
            ca_container[row['校区教练']]['high'] += 1
    
    # 转换为DataFrame并格式化比例
    ca_df = pd.DataFrame([{
        '校区教练': k,
        '5人及以下课程数': v['low'],
        '5人及以下比例': f"{(v['low']/v['total']*100):.2f}%" if v['total'] !=0 else "0.00%",
        '10人及以上课程数': v['high'],
        '10人及以上比例': f"{(v['high']/v['total']*100):.2f}%" if v['total'] !=0 else "0.00%"
    } for k, v in ca_container.items()])
    
    # 合并到主表（先删除可能存在的旧列，避免重复）
    target_cols = ['5人及以下课程数', '5人及以下比例', '10人及以上课程数', '10人及以上比例']
    combined = combined.drop(columns=[col for col in target_cols if col in combined.columns], errors='ignore')
    return pd.merge(combined, ca_df, on='校区教练', how='left').fillna({
        '5人及以下课程数': 0, '10人及以上课程数': 0,
        '5人及以下比例': "0.00%", '10人及以上比例': "0.00%"
    })


# --------------------------
# 子函数：7. 计算校区营收贡献比
# --------------------------
def calculate_campus_contribution(combined):
    """计算教练对本校区的营收贡献比"""
    # 计算各校区总营收
    campus_total = combined.groupby('校区', as_index=False)['确认收入'].sum().rename(columns={'确认收入': '校区总营收'})
    # 计算贡献比
    combined = pd.merge(combined, campus_total, on='校区', how='left')
    combined['本校区确收贡献比'] = combined.apply(
        lambda x: f"{(x['确认收入']/x['校区总营收']*100):.2f}%" if x['校区总营收'] !=0 else "0.00%", axis=1
    )
    return combined.drop(columns='校区总营收')  # 移除临时列，减少内存占用


# --------------------------
# 子函数：8. 统计校外教练校内上课节次
# --------------------------
def calculate_offcampus_in_campus(xnskjl, course_coach, off_campus_coaches):
    """统计校外教练在校内的上课节次"""
    xnskjl_copy = xnskjl.copy()
    xnskjl_copy['时加课'] = xnskjl_copy['上课信息'].astype(str) + xnskjl_copy['课程名称'].astype(str)
    # 去重课程并匹配教练
    xnskjl_unique = xnskjl_copy.drop_duplicates(subset=['时加课']).copy()
    xnskjl_unique = pd.merge(xnskjl_unique, course_coach, on='时加课', how='left')
    # 筛选校外教练并统计
    return xnskjl_unique[xnskjl_unique['教练'].isin(off_campus_coaches)].groupby(
        '教练', as_index=False
    ).size().rename(columns={'size': '校内上课节次'})


# --------------------------
# 子函数：9. 合并校外课时费数据
# --------------------------
def merge_teaching_fee(combined, xwksf):
    """合并校外课时费数据"""
    xwksf_copy = xwksf.copy()
    xwksf_copy['校区教练'] = xwksf_copy['校区'].astype(str) + '-' + xwksf_copy['教练'].astype(str)
    # 合并前删除旧列，避免冲突
    if '应发课时费' in combined.columns:
    # 存在则删除，errors='ignore' 进一步兜底（双重保险）
        combined = combined.drop(columns=['应发课时费'], errors='ignore')
    combined = pd.merge(combined, xwksf_copy[['校区教练', '应发课时费']], on='校区教练', how='left')
    combined['应发课时费'] = combined['应发课时费'].fillna(0).round(2)
    return combined

# --------------------------
# 子函数：10. 场地费以及相关费用计算
# --------------------------
def site_cost(combined,cdf,xwskjl,last_month_ana):
    # 读取Excel文件
    excel_file = pd.ExcelFile(cdf)
    
    # 存储所有sheet的数据
    sheet_data = []
    # print(excel_file.sheet_names)
    # 遍历所有sheet
    for sheet_name in excel_file.sheet_names:
        # 读取sheet数据，不设置表头（header=None）
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        # 删除第一行（标题行，如"N+体育农都校区场地使用签到表"）
        df = df.drop(index=0).reset_index(drop=True)
        df.columns = df.iloc[0]  # 将第一行设置为列名
        df = df.drop(index=0).reset_index(drop=True)  # 删除原来的第一行
        if sheet_name == '农都校区':
                df['校区'] = '农都城校区（篮球）'
        elif sheet_name == '滨江校区':
                df['校区'] = '滨江天街校区'
        elif sheet_name == '西溪校区':
                df['校区'] = '西溪天街校区'
         # 将当前sheet的数据追加到列表中
        sheet_data.append(df)
    # 合并所有sheet的数据
    date_useless,xwskjl_copy =calculate_basic_stats(xwskjl)
    # print(xwskjl_copy)
    date_sheet = pd.concat(sheet_data, ignore_index=True)
    date_sheet['上课日期'] = pd.to_datetime(date_sheet['上课日期'], errors='coerce').dt.strftime('%Y-%m-%d')
    date_sheet['上课时间'] = date_sheet['上课时间'].str.replace('--', '-', regex=False)
    date_sheet['时加课'] = date_sheet['上课日期'].astype(str) + ' ' + date_sheet['上课时间'].astype(str) + date_sheet['课程名称'].astype(str)
    capums_ana = date_sheet.groupby('校区',as_index=False)['场地费用'].sum().rename(columns={'场地费用':'校区场地费合计'})
    date_sheet = date_sheet.drop(['场地费用合计','场地编号' ,'周明细','合计'], axis=1, errors='ignore')
    xwskjl_in = xwskjl_copy[xwskjl_copy['状态'] == '已到'].copy()
    course_site_count = xwskjl_in.groupby('时加课', as_index=False).size().rename(columns={'size': '上课人数'})
    campus_attendance = xwskjl_in.groupby('校区',as_index=False).size().rename(columns={'size':'校区总到课人次'})
    # target_value = '滨江足球周六培优2班 25年秋'
    # # 3. 精确筛选（推荐用 loc，稳定不报错）
    # result = course_site_count[course_site_count['时加课'].str.contains(target_value, na=False)]
    # result1 = date_sheet[date_sheet['时加课'].str.contains(target_value, na=False)]
    # print(result)
    # print(result1['时加课'])
    # 2025-10-18 09:30-11:30滨江足球周六培优2班 25年秋
    # print(date_sheet)
    capums_ana  = pd.merge(capums_ana,campus_attendance , on='校区', how='left')
    date_sheet = pd.merge(date_sheet, course_site_count, on='时加课', how='left')
    capums_ana['校区生均场地费'] = capums_ana.apply(
        lambda x: round(x['校区场地费合计'] / x['校区总到课人次'], 2) if x['校区总到课人次'] != 0 else 0, axis=1
                )      
    # 删除不需要的列
    date_sheet['教练员'] = date_sheet['教练员'].str.replace(' ', '', regex=False)
    date_sheet['校区教练'] = date_sheet['校区'].astype(str) + '-' + date_sheet['教练员'].astype(str)
    df = date_sheet.groupby('校区教练',as_index=False)['场地费用'].sum().rename(columns={'场地费用':'场地费'})
    df['场地费'] = pd.to_numeric(df['场地费'], errors='coerce').fillna(0).round(2)
    combined = pd.merge(combined, df, on='校区教练', how='left')
    combined['S3:场地费生均成本'] = combined.apply(
        lambda x: round(x['场地费']/x['上课人次'],2) if x['上课人次'] !=0 else "0", axis=1)
    # print(combined.info())
    combined['(场地费+课时费)/确收'] = combined.apply(
        lambda x: f"{((x['场地费'] + x['应发课时费']) / x['确认收入'] * 100):.2f}%" if x['确认收入'] != 0 else "0", axis=1
    ) 
    date_sheet['课程生均场地费'] = date_sheet.apply(
        lambda x: round(x['场地费用']/x['上课人数'],2) if x['上课人数'] !=0 else 0, axis=1)
    date_sheet = pd.merge(
    date_sheet,
    capums_ana[['校区', '校区生均场地费']],
    on='校区',
    how='left'
    )
    # print(date_sheet[pd.isna(date_sheet['上课人数'])][['课程名称','上课日期','上课时间','上课人数','场地费用','教练员']])
    date_sheet_lower = date_sheet[date_sheet['课程生均场地费'] <= date_sheet['校区生均场地费']].groupby('校区教练',as_index=False)['课程生均场地费'].count().rename(columns={'课程生均场地费':'低于校区生均场地费占比'})
    # for i in date_sheet.items: 
    #      if i
    date_sheet_higher = date_sheet[date_sheet['课程生均场地费'] > date_sheet['校区生均场地费']*2].groupby('校区教练',as_index=False)['课程生均场地费'].count().rename(columns={'课程生均场地费':'高于校区生均场地费1倍占比'})
    date_sheet_highest  = date_sheet.groupby('校区教练',as_index=False)['课程生均场地费'].max().rename(columns={'课程生均场地费':'最高课程生均场地费'})
    combined = pd.merge(combined, date_sheet_lower, on='校区教练', how='left').fillna({'低于校区生均场地费占比':0})
    combined = pd.merge(combined, date_sheet_higher, on='校区教练', how='left').fillna({'高于校区生均场地费1倍占比':0})
    combined = pd.merge(combined, date_sheet_highest, on='校区教练', how='left').fillna({'最高课程生均场地费':0})

    combined['低于校区生均场地费占比'] = combined.apply(
        lambda x: f"{(x['低于校区生均场地费占比']/x['课次']*100):.2f}%" if x['课次'] !=0 else "0", axis=1
                )
    combined['高于校区生均场地费1倍占比'] = combined.apply(
        lambda x: f"{(x['高于校区生均场地费1倍占比']/x['课次']*100):.2f}%" if x['课次'] !=0 else "0", axis=1
                )
    combined['最高课程生均场地费/s3'] = combined.apply(
        lambda x: f"{(x['最高课程生均场地费']/x['S3:场地费生均成本']):.2f}倍", axis=1)
    
    last_month_ana['校区'] = last_month_ana['校区'].ffill()
    last_month_ana['校区'] = last_month_ana['校区'].replace('农都城校区', '农都城校区（篮球）')
    last_month_ana['校区教练'] = last_month_ana['校区'].astype(str) + '-' + last_month_ana['教练'].astype(str)
    combined = pd.merge(combined, last_month_ana[['校区教练','S3:场地费生均成本']], on='校区教练', how='left').rename(columns={'S3:场地费生均成本_y':'上月场地费生均成本','S3:场地费生均成本_x':'S3:场地费生均成本'})
    combined['上月场地费生均成本'] = combined['上月场地费生均成本'].round(2)
    combined['s3较上月变化'] = combined.apply(
         lambda x: f"{((x['S3:场地费生均成本']-x['上月场地费生均成本'])/x['上月场地费生均成本']*100):.2f}%" if x['上月场地费生均成本'] !=0 else "0", axis=1
                )
    return combined
    # print(combined[['校区教练','S3:场地费生均成本','上月场地费生均成本']])
    # print(combined['(场地费+课时费)/确收'])

    # return date_sheet

# --------------------------
# 子函数：11. 安全导出Excel并整理列顺序
# --------------------------
def export_result(combined_data, df, export_dir):
    """
    导出两个 DataFrame 到 Excel，自动命名（按上个月），避免覆盖，并应用格式。
    
    :param combined_data: 主数据表
    :param df: 完整数据表
    :param export_dir: 用户指定的导出目录（字符串）
    :return: 实际保存的主文件路径
    """
    # 确保导出目录存在
    os.makedirs(export_dir, exist_ok=True)

    # 按上个月生成文件名
    current = datetime.datetime.now()
    prev_month = current - relativedelta(months=1)
    base_name = f"{prev_month.year}年{prev_month.month}月校外分析.xlsx"

    # 构建初始路径
    export_path = os.path.join(export_dir, base_name)
    export_path_df = os.path.join(export_dir, f"{prev_month.year}年{prev_month.month}月校外分析_完整数据.xlsx")

    # 避免覆盖：同步递增序号
    counter = 1
    while os.path.exists(export_path) or os.path.exists(export_path_df):
        export_path = os.path.join(export_dir, f"{prev_month.year}年{prev_month.month}月校外分析_{counter}.xlsx")
        export_path_df = os.path.join(export_dir, f"{prev_month.year}年{prev_month.month}月校外分析_{counter}_完整数据.xlsx")
        counter += 1

    # ---------------------- 核心：带格式的导出函数 ----------------------
    def export_with_format(dataframe, output_path, sheet_name='综合统计'):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

        # 写入数据
        for r in dataframe_to_rows(dataframe, index=False, header=True):
            ws.append(r)

        # 居中对齐
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_alignment

        # 自适应列宽（中文友好）
        def calculate_char_width(text):
            if text is None:
                return 0
            width = 0
            for char in str(text):
                width += 2 if '\u4e00' <= char <= '\u9fff' else 1
            return width

        for column in ws.columns:
            max_width = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    w = calculate_char_width(cell.value)
                    if w > max_width:
                        max_width = w
                except:
                    pass
            adjusted_width = min(max_width + 3, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)

    # ---------------------- 执行导出 ----------------------
    try:
        export_with_format(combined_data, export_path, '综合统计')
        export_with_format(df, export_path_df, '综合统计')
        print(f"✅ 主文件已保存: {export_path}")
        print(f"✅ 完整数据已保存: {export_path_df}")
        return export_path

    except PermissionError:
        print("❌ 导出失败：目标文件被占用（请关闭已打开的同名Excel文件）")
        # 使用 export_dir 作为备份目录（不再用 current_folder）
        backup_suffix = datetime.datetime.now().strftime("%H%M%S")
        backup_path = os.path.join(export_dir, f"{prev_month.year}年{prev_month.month}月校外分析_备份_{backup_suffix}.xlsx")
        
        # 备份两个 sheet 到一个文件
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '综合统计'
        for r in dataframe_to_rows(combined_data, index=False, header=True):
            ws1.append(r)
        
        ws2 = wb.create_sheet('综合统计_完整数据')
        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)
        
        # 应用格式（复用逻辑）
        center_alignment = Alignment(horizontal='center', vertical='center')
        def calculate_char_width(text):
            if text is None:
                return 0
            width = 0
            for char in str(text):
                width += 2 if '\u4e00' <= char <= '\u9fff' else 1
            return width
        
        for ws in [ws1, ws2]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment
            for col in ws.columns:
                max_w = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        w = calculate_char_width(cell.value)
                        if w > max_w:
                            max_w = w
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_w + 3, 60)
        
        wb.save(backup_path)
        print(f"📁 已备份到: {backup_path}")
        return backup_path

    except Exception as e:
        print(f"❌ 导出发生未知错误: {str(e)}")
        backup_suffix = datetime.datetime.now().strftime("%H%M%S")
        backup_path = os.path.join(export_dir, f"{prev_month.year}年{prev_month.month}月校外分析_备份_{backup_suffix}.xlsx")
        
        # 同样备份到 export_dir
        wb = Workbook()
        ws1 = wb.active
        ws1.title = '综合统计'
        for r in dataframe_to_rows(combined_data, index=False, header=True):
            ws1.append(r)
        ws2 = wb.create_sheet('综合统计_完整数据')
        for r in dataframe_to_rows(df, index=False, header=True):
            ws2.append(r)
        
        # 简化格式（避免再次出错）
        for ws in [ws1, ws2]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(backup_path)
        print(f"📁 已备份（无格式）到: {backup_path}")
        return backup_path
    
# --------------------------
# 子函数：12. 校区汇总
# --------------------------
def summarize_campus(group,xwskjl,last_month_ana):
    total_classes = int(group['课次'].sum())  # 课次：整数
    total_students = int(group['上课人次'].sum())  # 上课人次：整数
    avg_class_size = round(total_students / total_classes, 2) if total_classes > 0 else 0.00  # 保留2位小数
    print(total_students)
    # 到课率：修复逻辑（按当前分组的校区筛选数据，确保对齐）
    current_campus = group['校区'].iloc[0]  # 获取当前分组的校区名称
    # 筛选当前校区的应到数据，按校区汇总总应到人次
    campus_attedance_total = xwskjl[xwskjl['校区'] == current_campus].shape[0]  # 当前校区总应到人次
    print(campus_attedance_total)
    # 计算到课率（当前校区已到人数 ÷ 当前校区总应到人数）
    if campus_attedance_total > 0:
        attendance_rate = total_students / campus_attedance_total
        avg_attendance = f"{attendance_rate*100:.2f}%"
    else:
        avg_attendance = "0.00%"
    
    # 其他字段计算逻辑不变
    total_revenue = round(group['确认收入'].sum(), 2)  # 确认收入：2位小数
    avg_price = round(total_revenue / total_students, 2) if total_students > 0 else 0.00  # 平均课程单价：2位小数
    total_small_classes = int(group['5人及以下课程数'].sum())  # 整数
    small_class_ratio = f"{(total_small_classes / total_classes * 100):.2f}%" if total_classes > 0 else "0.00%"
    total_large_classes = int(group['10人及以上课程数'].sum())  # 整数
    large_class_ratio = f"{(total_large_classes / total_classes * 100):.2f}%" if total_classes > 0 else "0.00%"
    total_internal_classes = int(group['校内上课节次'].sum())  # 整数
    total_teacher_fee = round(group['应发课时费'].sum(), 2)  # 应发课时费：1位小数
    total_venue_fee = round(group['场地费'].sum(), 1)  # 场地费：1位小数
    avg_venue_cost_per_student = round(total_venue_fee / total_students, 2) if total_students > 0 else 0.00  # 2位小数
    cost_revenue_ratio = f"{((total_venue_fee + total_teacher_fee) / total_revenue * 100):.2f}%" if total_revenue > 0 else "0.00%"
    last_month_campus_data = last_month_ana[last_month_ana['校区'] == current_campus]
    print(cost_revenue_ratio)
    if not last_month_campus_data.empty:
        total_venue_last = last_month_campus_data['场地费'].sum()
        total_students_last = last_month_campus_data['上课人次'].sum()
        if total_students_last > 0:
            last_month_avg_venue_cost_per_student = round(total_venue_last / total_students_last, 2)
        else:
            last_month_avg_venue_cost_per_student = 0.0
    else:
        last_month_avg_venue_cost_per_student = 0.0

    # 计算变化率
    if abs(last_month_avg_venue_cost_per_student) < 1e-6:
        s3_change = "0.00%"
    else:
        s3_change = f"{((avg_venue_cost_per_student - last_month_avg_venue_cost_per_student) / last_month_avg_venue_cost_per_student * 100):.2f}%"
    print(cost_revenue_ratio)

    return pd.Series({
        '校区': f"{current_campus}-汇总",
        '课次': total_classes,
        '上课人次': total_students,
        '平均课堂人次': avg_class_size,
        '到课率': avg_attendance,
        '确认收入': total_revenue,
        '平均课程单价': avg_price,
        '5人及以下课程数': total_small_classes,
        '5人及以下比例': small_class_ratio,
        '10人及以上课程数': total_large_classes,
        '10人及以上比例': large_class_ratio,
        '校内上课节次': total_internal_classes,
        '应发课时费': total_teacher_fee,
        '场地费': total_venue_fee,
        'S3:场地费生均成本': avg_venue_cost_per_student,
        '上月场地费生均成本': last_month_avg_venue_cost_per_student,  # 现在是正确值
        '(场地费+课时费)/确收': cost_revenue_ratio,
        's3较上月变化': s3_change
    }, dtype=object)


# --------------------------
# 子函数：13. 全部汇总
# --------------------------
def summarize_total(combined,xwskjl,last_month_ana):
    """计算所有校区的总汇总数据"""
    # 筛选非汇总行（避免重复计算校区汇总行）
    combined = combined[~combined['校区'].str.contains('乒乓球')].copy()
    non_summary_data = combined[~combined['校区'].str.endswith('-汇总', na=False)].copy()
    
    total_classes = int(non_summary_data['课次'].sum())  # 总课次
    total_students = int(non_summary_data['上课人次'].sum())  # 总上课人次
    print(total_students)
    avg_class_size = round(total_students / total_classes, 2) if total_classes > 0 else 0.00  # 总平均课堂人次
    # 总到课率（总已到人次 ÷ 总应到人次）
    xwskjl_non_pingpong = xwskjl[~xwskjl['校区'].str.contains('乒乓球')].copy()
    total_attendance_possible = xwskjl_non_pingpong.shape[0] # 总已到人次（复用上课人次）
    print(total_attendance_possible)
    # 若需精准应到人次，需从原始xwskjl计算（此处沿用现有逻辑，与校区汇总一致）
    total_attendance_rate = f"{(total_students / total_attendance_possible * 100):.2f}%" if total_attendance_possible > 0 else "0.00%"
    
    # 营收相关
    total_revenue = round(non_summary_data['确认收入'].sum(), 2)  # 总确认收入
    avg_price = round(total_revenue / total_students, 2) if total_students > 0 else 0.00  # 总平均课程单价
    
    # 班级规模相关
    total_small_classes = int(non_summary_data['5人及以下课程数'].sum())  # 总5人及以下课程数
    small_class_ratio = f"{(total_small_classes / total_classes * 100):.2f}%" if total_classes > 0 else "0.00%"
    total_large_classes = int(non_summary_data['10人及以上课程数'].sum())  # 总10人及以上课程数
    large_class_ratio = f"{(total_large_classes / total_classes * 100):.2f}%" if total_classes > 0 else "0.00%"
    
    # 成本相关
    total_internal_classes = int(non_summary_data['校内上课节次'].sum())  # 总校内上课节次
    total_teacher_fee = round(non_summary_data['应发课时费'].sum(), 2)  # 总应发课时费
    total_venue_fee = round(non_summary_data['场地费'].sum(), 1)  # 总场地费
    avg_venue_cost_per_student = round(total_venue_fee / total_students, 2) if total_students > 0 else 0.00  # 总场地费生均成本
    cost_revenue_ratio = f"{((total_venue_fee + total_teacher_fee) / total_revenue * 100):.2f}%" if total_revenue > 0 else "0.00%"
    
    #获取上月数据
    
    
    last_month_data = last_month_ana[last_month_ana['校区'] == 'Z总计']
    if not last_month_data.empty:
        last_month_average_venue_fee = last_month_data['S3:场地费生均成本'].values[0]
        last_month_average_venue_fee = round(last_month_average_venue_fee, 2)
        s3_change = f"{((avg_venue_cost_per_student - last_month_average_venue_fee)/last_month_average_venue_fee*100):.2f}%" if last_month_average_venue_fee!= 0 else "0.00%"
    else:
        s3_change = "0.00%"
    return pd.Series({
        '校区': 'Z总计',  # 汇总标识
        # '教练': '',  # 总汇总无特定教练
        '课次': total_classes,
        '上课人次': total_students,
        '平均课堂人次': avg_class_size,
        '到课率': total_attendance_rate,
        '确认收入': total_revenue,
        '平均课程单价': avg_price,
        '5人及以下课程数': total_small_classes,
        '5人及以下比例': small_class_ratio,
        '10人及以上课程数': total_large_classes,
        '10人及以上比例': large_class_ratio,
        '校内上课节次': total_internal_classes,
        '应发课时费': total_teacher_fee,
        '场地费': total_venue_fee,
        'S3:场地费生均成本': avg_venue_cost_per_student,
        '上月场地费生均成本': last_month_average_venue_fee,
        '本校区营收贡献比': '-',  # 总汇总无校区贡献比
        '(场地费+课时费)/确收': cost_revenue_ratio,
        's3较上月变化': s3_change,
        '低于校区生均场地费占比': '-',  # 总汇总无需该维度
        '高于校区生均场地费1倍占比': '-',  # 总汇总无需该维度
        '最高课程生均场地费/s3': '-'  # 总汇总无需该维度
    }, dtype=object)

#  --------------------------
# 总函数：串联所有子函数，执行完整分析
# --------------------------
def coach_comprehensive_analysis(xwskjl, xnskjl, xwksf, cw, off_campus_coaches,last_month_ana,export_dir):
    """总函数：调用所有子函数，返回最终结果和导出路径"""
    # 1. 生成教练归属表（当前逻辑中未直接使用，保留以备扩展）
    # jl = create_coach_belonging(cw, off_campus_coaches)
    
    # 2. 处理财务数据
    xwcw, course_coach = process_finance_data(cw)
    
    # 3. 计算基础统计
    combined, xwskjl_copy = calculate_basic_stats(xwskjl)
    
    # 4. 计算到课率
    combined = calculate_attendance_rate(combined, xwskjl_copy)
    
    # 5. 统计收入与单价
    combined = calculate_revenue(combined, xwcw)
    
    # 6. 分析班级规模
    combined = analyze_class_size(combined, xwskjl_copy)
    
    # 7. 计算校区营收贡献比
    combined = calculate_campus_contribution(combined)
    
    # 8. 统计校外教练校内课次
    xn_skjl = calculate_offcampus_in_campus(xnskjl, course_coach, off_campus_coaches)
    combined = pd.merge(combined, xn_skjl, on='教练', how='left')
    combined['校内上课节次'] = combined['校内上课节次'].fillna(0).astype(int)
    
    # 9. 合并课时费
    combined = merge_teaching_fee(combined, xwksf)
    
    # 10. 场地费计算
    combined  = site_cost(combined,cdf,xwskjl,last_month_ana)
    combined = combined[~combined['校区'].str.contains('乒乓球')].copy()

    
    # 11. 添加校区汇总行
    summary_rows = combined.groupby('校区', group_keys=False).apply(
    lambda group: summarize_campus(group, xwskjl_copy,last_month_ana)  # 正确传递xwskjl_copy参数
    ).reset_index(drop=True).fillna(' ')
    
    # print(tabulate(summary_rows, headers='keys', tablefmt='psql', showindex=False))
    combined = pd.concat([combined, summary_rows], ignore_index=True)
    # 12. 添加全部汇总行
    total_summary = summarize_total(combined,xwskjl,last_month_ana).to_frame().T  # 转换为DataFrame
    combined = pd.concat([combined, total_summary], ignore_index=True)  # 拼接总汇总行
    
    # 13. 整理列顺序并导出
    final_cols = [
        '校区', '教练',  '课次', '上课人次', '平均课堂人次', '到课率',
        '确认收入', '平均课程单价','场地费','应发课时费','S3:场地费生均成本','(场地费+课时费)/确收',
        '本校区营收贡献比','校内上课节次','5人及以下课程数', '5人及以下比例',
        '10人及以上课程数', '10人及以上比例','上月场地费生均成本',
        's3较上月变化','低于校区生均场地费占比','高于校区生均场地费1倍占比','最高课程生均场地费/s3'
    ]
    df = combined.copy()
    combined = combined[final_cols].sort_values(['校区','确认收入'], ascending=[True,True]).reset_index(drop=True)
    
    # 13. 导出结果
    export_path = export_result(combined,df,export_dir)
    
    # return combined, df
    return df,combined,export_path


# --------------------------
# 调用示例
# --------------------------
if __name__ == "__main__":
    # 1. 读取数据（替换为实际路径）
    # xwskjl = pd.read_excel('/Users/wangjun/Desktop/华越分析/python分析/9月校外上课记录.xlsx', sheet_name='Sheet1')
    # xnskjl = pd.read_excel('/Users/wangjun/Desktop/华越分析/python分析/9月校内上课记录.xlsx', sheet_name='Sheet1')
    skjl = pd.read_excel(r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\3月\校外\3月上课记录.xls',sheet_name="Sheet1")
    xwksf = pd.read_excel(r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\3月\校外\202603 校外课时费.xlsx', sheet_name='Sheet1')
    cw = pd.read_excel(r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\3月\校外\3月财务数据.xls', sheet_name='财务')
    cdf = r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\3月\校外\N+体育3月场地使用费用汇总表.xlsx'
    export_dir = r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\3月\校外'
    last_month_ana = pd.read_excel(r'C:\Users\Administrator\Desktop\python\代码分析\华越分析\华越分析\2026年分析\2月\校外\2026年2月校外分析.xlsx', sheet_name='综合统计')
    # last_month_ana = pd.read_excel()
    # 2. 定义校外教练列表
    off_campus_coaches = ["邓文凯", "唐崇尧", "王峥", "卞梦男", "苏长锴", "甘洪宇", "何苗", "邱毅俊", "邵志龙", "钱威", "徐思澳",'徐子杰','许佳晴']
    xwskjl, xnskjl = skjl_sperate(skjl,cw)
    # 3. 执行分析
    self_ana_result,final_result,exprot_path = coach_comprehensive_analysis(
        xwskjl=xwskjl, xnskjl=xnskjl, xwksf=xwksf, cw=cw, off_campus_coaches=off_campus_coaches,last_month_ana = last_month_ana,export_dir=export_dir
    )
    # coach_comprehensive_analysis(
    #     xwskjl=xwskjl, xnskjl=xnskjl, xwksf=xwksf, cw=cw, off_campus_coaches=off_campus_coaches,last_month_ana = last_month_ana
    # )
    print('文件已导出')
